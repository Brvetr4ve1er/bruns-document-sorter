"""
Export Page — Phase 3
CSV and Excel export with column selection, live preview, and filtered export.
"""
import io
import os
import sqlite3
from datetime import datetime

import streamlit as st
import pandas as pd

from ui.styles import COLORS


# Friendly display names → actual DB column names (from fetch_containers_view query)
EXPORT_COLUMNS = {
    "N° Container":        "N° Container",
    "N° TAN":              "N° TAN",
    "Item / Description":  "Item",
    "Compagnie maritime":  "Compagnie maritime",
    "Port":                "Port",
    "Transitaire":         "Transitaire",
    "ETD (Date shipment)": "Date shipment",
    "ETA (Date accostage)":"Date accostage",
    "Statut Container":    "Statut Container",
    "Container size":      "Container size",
    "N° Seal":             "N° Seal",
    "Date livraison":      "Date livraison",
    "Site livraison":      "Site livraison",
    "Chauffeur livraison": "Chauffeur livraison",
    "Date dépotement":     "Date dépotement",
    "Date restitution":    "Date restitution",
    "Chauffeur restitution":"Chauffeur restitution",
    "Commentaire":         "Commentaire",
    "Source file":         "Source",
    "Créé le":             "Créé le",
}


def render_export_page():
    """Full export dashboard page."""

    st.markdown("""
    <div data-aos="fade-up" style="margin-bottom: 8px;">
        <p style="color:#8C95B3; margin:0; font-size:0.95rem;">
        Export your logistics data to CSV or Excel. Select columns, apply filters, and download instantly.
        </p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")

    # ── Load data ──────────────────────────────────────────────────────────────
    try:
        df = _fetch_all_data()
    except Exception as e:
        st.error(f"❌ Could not load data: {e}")
        return

    if df.empty:
        st.info("📭 No data to export. Process some PDF documents first.")
        return

    # ── Column selection ───────────────────────────────────────────────────────
    st.markdown("#### 📋 Select Columns to Export")

    col1, col2, col3 = st.columns([1, 1, 1])
    with col1:
        if st.button("✅ Select All", key="select_all_cols"):
            st.session_state["export_selected_cols"] = list(EXPORT_COLUMNS.keys())
    with col2:
        if st.button("🗑️ Clear All", key="clear_all_cols"):
            st.session_state["export_selected_cols"] = []
    with col3:
        if st.button("📌 Core Fields Only", key="core_cols"):
            st.session_state["export_selected_cols"] = [
                "N° Container", "N° TAN", "Item / Description",
                "Compagnie maritime", "Statut Container",
                "ETD (Date shipment)", "ETA (Date accostage)"
            ]

    default_cols = st.session_state.get(
        "export_selected_cols", list(EXPORT_COLUMNS.keys())
    )
    selected_labels = st.multiselect(
        "Columns",
        options=list(EXPORT_COLUMNS.keys()),
        default=[c for c in default_cols if c in EXPORT_COLUMNS],
        key="export_cols_select",
        help="Choose which columns to include in the export"
    )
    st.session_state["export_selected_cols"] = selected_labels

    if not selected_labels:
        st.warning("⚠️ Please select at least one column.")
        return

    # ── Quick status filter for export ────────────────────────────────────────
    st.markdown("#### 🔽 Filter Export Data")
    fcol1, fcol2, fcol3 = st.columns(3)

    with fcol1:
        status_opts = sorted([x for x in df["Statut Container"].unique() if pd.notna(x)])
        export_status = st.multiselect(
            "Status filter",
            options=status_opts,
            default=[],
            key="export_status_filter"
        )
    with fcol2:
        carrier_opts = sorted([x for x in df["Compagnie maritime"].unique() if pd.notna(x)])
        export_carrier = st.multiselect(
            "Carrier filter",
            options=carrier_opts,
            default=[],
            key="export_carrier_filter"
        )
    with fcol3:
        search_export = st.text_input(
            "Text search",
            value="",
            key="export_search",
            placeholder="Container, TAN, Item..."
        )

    # Apply filters to export df
    export_df = df.copy()
    if export_status:
        export_df = export_df[export_df["Statut Container"].isin(export_status)]
    if export_carrier:
        export_df = export_df[export_df["Compagnie maritime"].isin(export_carrier)]
    if search_export:
        term = search_export.lower()
        mask = (
            export_df["N° Container"].astype(str).str.lower().str.contains(term, na=False) |
            export_df["N° TAN"].astype(str).str.lower().str.contains(term, na=False) |
            export_df["Item"].astype(str).str.lower().str.contains(term, na=False)
        )
        export_df = export_df[mask]

    # ── Preview ────────────────────────────────────────────────────────────────
    # Map selected labels to actual column names
    actual_cols = [EXPORT_COLUMNS[lbl] for lbl in selected_labels if EXPORT_COLUMNS[lbl] in export_df.columns]
    preview_df = export_df[actual_cols].head(10)

    st.markdown(f"#### 👁️ Preview ({min(10, len(export_df))} of {len(export_df):,} rows)")
    st.dataframe(preview_df, use_container_width=True, hide_index=True)

    st.markdown("---")

    # ── Download buttons ───────────────────────────────────────────────────────
    st.markdown("#### 📥 Download")

    full_export = export_df[actual_cols]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    dcol1, dcol2, dcol3 = st.columns(3)

    with dcol1:
        csv_data = full_export.to_csv(index=False, encoding="utf-8-sig")
        st.download_button(
            label=f"📄 Download CSV ({len(full_export):,} rows)",
            data=csv_data.encode("utf-8-sig"),
            file_name=f"bruns_logistics_{timestamp}.csv",
            mime="text/csv",
            use_container_width=True,
            type="primary"
        )

    with dcol2:
        excel_data = _to_excel(full_export)
        st.download_button(
            label=f"📊 Download Excel ({len(full_export):,} rows)",
            data=excel_data,
            file_name=f"bruns_logistics_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    with dcol3:
        # Export ALL data regardless of filters
        all_cols_df = df[actual_cols]
        csv_all = all_cols_df.to_csv(index=False, encoding="utf-8-sig")
        st.download_button(
            label=f"📦 Export ALL ({len(df):,} rows, no filter)",
            data=csv_all.encode("utf-8-sig"),
            file_name=f"bruns_logistics_ALL_{timestamp}.csv",
            mime="text/csv",
            use_container_width=True
        )

    # ── Stats ─────────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 📊 Database Summary")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total rows", f"{len(df):,}")
    c2.metric("Filtered rows", f"{len(export_df):,}")
    c3.metric("Columns selected", f"{len(actual_cols)}")
    status_counts = df["Statut Container"].value_counts()
    c4.metric("Status types", f"{len(status_counts)}")

    with st.expander("📈 Status Breakdown"):
        st.bar_chart(status_counts)


def _fetch_all_data() -> pd.DataFrame:
    """Fetch complete data for export (mirrors fetch_containers_view in app_new.py)."""
    import config
    conn = sqlite3.connect(config.DB_PATH)
    conn.row_factory = sqlite3.Row
    df = pd.read_sql_query("""
        SELECT
            c.id AS container_id,
            s.id AS shipment_id,
            c.container_number   AS "N° Container",
            s.tan                AS "N° TAN",
            s.item_description   AS "Item",
            s.compagnie_maritime AS "Compagnie maritime",
            s.port               AS "Port",
            s.transitaire        AS "Transitaire",
            s.etd                AS "Date shipment",
            s.eta                AS "Date accostage",
            c.statut_container   AS "Statut Container",
            c.size               AS "Container size",
            c.seal_number        AS "N° Seal",
            c.date_livraison     AS "Date livraison",
            c.site_livraison     AS "Site livraison",
            c.date_depotement    AS "Date dépotement",
            c.date_restitution   AS "Date restitution",
            c.restitue_chauffeur AS "Chauffeur restitution",
            c.livre_chauffeur    AS "Chauffeur livraison",
            c.commentaire        AS "Commentaire",
            s.source_file        AS "Source",
            c.created_at         AS "Créé le"
        FROM containers c
        JOIN shipments s ON s.id = c.shipment_id
        ORDER BY c.id DESC
    """, conn)
    conn.close()
    return df


def _to_excel(df: pd.DataFrame) -> bytes:
    """Convert dataframe to formatted Excel bytes."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Logistics Data")
        ws = writer.sheets["Logistics Data"]

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=10
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Style header row
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color="1B2341", end_color="1B2341", fill_type="solid")
        header_font = Font(color="E6E9F2", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    return output.getvalue()
