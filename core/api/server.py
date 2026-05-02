"""
core/api/server.py

Live REST API + HTML UI server.

API endpoints (port 7845, JSON, for Power BI / external):
    GET /api/status
    GET /api/logistics/shipments?page=1&page_size=50
    GET /api/logistics/containers?...
    GET /api/logistics/shipments_full?...
    GET /api/travel/persons|families|documents

HTML UI routes (same port, server-rendered + HTMX):
    GET /                  → mode picker (splash)
    GET /logistics         → logistics dashboard
    GET /travel            → travel dashboard (P4 — placeholder for now)

Start with:
    python -m core.api.server
    or
    python core/api/server.py
"""

import json
import logging
import os
import sys

from flask import Flask, jsonify, request, abort, render_template, send_file
from flask_cors import CORS

# Allow running from root
ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

# TD12: configure structured logging before any imports that emit log records.
from core.logging_config import configure_logging
configure_logging()
log = logging.getLogger(__name__)

from core.storage.paginator import paginated_query
from core.storage.db import get_connection

# ─── Bootstrap: register every domain's LLM prompts at server startup ─────────
# Without this, the extraction pipeline raises
#   "No prompt registered for module 'X' and doc_type 'UNKNOWN'"
# at the first upload. The modules define init_prompts() but don't auto-call it.
try:
    from modules.logistics.prompts import init_prompts as _init_logistics_prompts
    _init_logistics_prompts()
except Exception as _e:
    log.warning("failed to register logistics prompts: %s", _e)
try:
    from modules.travel.prompts import init_prompts as _init_travel_prompts
    _init_travel_prompts()
except Exception as _e:
    log.warning("failed to register travel prompts: %s", _e)

app = Flask(__name__)
app.config["JSON_AS_ASCII"] = False   # serve French chars as UTF-8, not \uXXXX escapes

# TD9 fix: restrict CORS to localhost origins. Power BI Desktop hits the
# REST endpoints from the same machine, so localhost / 127.0.0.1 covers the
# real-world use case. If a customer needs to read these endpoints from
# another LAN machine, they can override via the BRUNS_CORS_ORIGINS env var
# (comma-separated list of explicit origins, e.g. "http://10.0.0.5:7845").
#
# flask_cors accepts a list of regex pattern strings; it does NOT correctly
# filter when given compiled re.Pattern objects mixed with strings, so we
# stay with strings.
_default_cors_origins = [
    r"http://localhost(:\d+)?",
    r"http://127\.0\.0\.1(:\d+)?",
]
_cors_env = os.environ.get("BRUNS_CORS_ORIGINS", "").strip()
if _cors_env:
    _cors_origins: list = [o.strip() for o in _cors_env.split(",") if o.strip()]
else:
    _cors_origins = _default_cors_origins
CORS(app, resources={r"/api/*": {"origins": _cors_origins}})

# Both databases live at <repo>/data/ — one canonical location.
DATA_DIR     = os.environ.get("BRUNS_DATA_DIR", os.path.join(ROOT_DIR, "data"))
LOGISTICS_DB = os.environ.get("BRUNS_LOGISTICS_DB",
                              os.path.join(DATA_DIR, "logistics.db"))
TRAVEL_DB    = os.environ.get("BRUNS_TRAVEL_DB",
                              os.path.join(DATA_DIR, "travel.db"))

PAGE_SIZE_MAX  = 200   # hard cap — prevents BI tools from dumping the whole DB in one shot


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _page_params():
    try:
        page      = max(1, int(request.args.get("page", 1)))
        page_size = min(PAGE_SIZE_MAX, max(1, int(request.args.get("page_size", 50))))
    except (ValueError, TypeError):
        abort(400, "page and page_size must be integers")
    return page, page_size

def _rows_to_dicts(rows: list) -> list[dict]:
    """Convert sqlite3.Row objects to plain JSON-serializable dicts."""
    return [dict(r) for r in rows]

def _paginated_response(conn, table, page, page_size, where="", order_by="id DESC", params=()):
    result = paginated_query(conn, table, page, page_size, where, order_by, params)
    return jsonify({
        "data":        _rows_to_dicts(result["rows"]),
        "pagination": {
            "page":        result["page"],
            "page_size":   result["page_size"],
            "total":       result["total"],
            "total_pages": result["total_pages"],
            "has_next":    result["has_next"],
            "has_prev":    result["has_prev"],
        }
    })


# ─── System Health ─────────────────────────────────────────────────────────────

@app.route("/api/status")
def api_status():
    checks = {}
    for label, db_path in [("logistics", LOGISTICS_DB), ("travel", TRAVEL_DB)]:
        checks[label] = "ok" if os.path.exists(db_path) else "missing"
    return jsonify({"status": "online", "databases": checks})


# ─── Logistics Endpoints ───────────────────────────────────────────────────────

@app.route("/api/logistics/shipments")
def logistics_shipments():
    if not os.path.exists(LOGISTICS_DB):
        return jsonify({"data": [], "pagination": {}}), 200
    page, page_size = _page_params()
    status = request.args.get("status")          # optional filter
    where  = "status = ?" if status else ""
    params = (status,)      if status else ()
    conn   = get_connection(LOGISTICS_DB)
    try:
        return _paginated_response(conn, "shipments", page, page_size, where, "id DESC", params)
    finally:
        conn.close()


@app.route("/api/logistics/containers")
def logistics_containers():
    if not os.path.exists(LOGISTICS_DB):
        return jsonify({"data": [], "pagination": {}}), 200
    page, page_size = _page_params()
    status = request.args.get("status")
    where  = "statut_container = ?" if status else ""
    params = (status,)              if status else ()
    conn   = get_connection(LOGISTICS_DB)
    try:
        return _paginated_response(conn, "containers", page, page_size, where, "id DESC", params)
    finally:
        conn.close()


# ─── Logistics: Flat denormalized view (one row per container) ─────────────────
# This is the primary endpoint for Power BI — no joins needed on the BI side.

@app.route("/api/logistics/shipments_full")
def logistics_shipments_full():
    """Flat JOIN of shipments + containers — one row per container, all columns."""
    if not os.path.exists(LOGISTICS_DB):
        return jsonify({"data": [], "pagination": {}}), 200

    page, page_size = _page_params()
    offset = (page - 1) * page_size

    conn = get_connection(LOGISTICS_DB)
    try:
        # Optional filters
        status    = request.args.get("status")
        carrier   = request.args.get("carrier")
        tan       = request.args.get("tan")

        where_clauses = []
        params = []
        if status:
            where_clauses.append("s.status = ?")
            params.append(status)
        if carrier:
            where_clauses.append("s.compagnie_maritime = ?")
            params.append(carrier)
        if tan:
            where_clauses.append("s.tan LIKE ?")
            params.append(f"%{tan}%")

        where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

        total = conn.execute(
            f"SELECT COUNT(*) FROM containers c JOIN shipments s ON s.id = c.shipment_id {where_sql}",
            params
        ).fetchone()[0]

        rows = conn.execute(
            f"""
            SELECT
                s.id                    AS shipment_id,
                c.id                    AS container_id,
                c.container_number      AS "N° Container",
                s.tan                   AS "N° TAN",
                s.item_description      AS "Item",
                s.compagnie_maritime    AS "Compagnie maritime",
                s.port                  AS "Port",
                s.transitaire           AS "Transitaire",
                s.vessel                AS "Navire",
                s.etd                   AS "Date shipment",
                s.eta                   AS "Date accostage",
                s.status                AS "Statut Expédition",
                s.document_type         AS "Type document",
                c.statut_container      AS "Statut Container",
                c.size                  AS "Container size",
                c.seal_number           AS "N° Seal",
                c.date_livraison        AS "Date livraison",
                c.site_livraison        AS "Site livraison",
                c.date_depotement       AS "Date dépotement",
                c.date_debut_surestarie AS "Date début Surestarie",
                c.date_restitution_estimative AS "Date restitution estimative",
                c.nbr_jours_surestarie_estimes AS "Nbr jours surestarie estimés",
                c.nbr_jours_perdu_douane       AS "Nbr jours perdu en douane",
                c.date_restitution      AS "Date réstitution",
                c.restitue_camion       AS "Réstitué par (Camion)",
                c.restitue_chauffeur    AS "Réstitué par (Chauffeur)",
                c.centre_restitution    AS "Centre de réstitution",
                c.livre_camion          AS "Livré par (Camion)",
                c.livre_chauffeur       AS "Livré par (Chauffeur)",
                c.montant_facture_check AS "Montant facturé (check)",
                c.nbr_jour_surestarie_facture AS "Nbr jour surestarie Facturé",
                c.montant_facture_da    AS "Montant facturé (DA)",
                c.taux_de_change        AS "Taux de change",
                c.n_facture_cm          AS "N° Facture compagnie maritime",
                c.commentaire           AS "Commentaire",
                c.date_declaration_douane   AS "Date declaration douane",
                c.date_liberation_douane    AS "Date liberation douane",
                s.source_file           AS "Source",
                c.created_at            AS "Créé le",
                c.modified_at           AS "Modifié le"
            FROM containers c
            JOIN shipments s ON s.id = c.shipment_id
            {where_sql}
            ORDER BY c.id DESC
            LIMIT ? OFFSET ?
            """,
            params + [page_size, offset]
        ).fetchall()

        total_pages = max(1, -(-total // page_size))  # ceiling division
        return jsonify({
            "data": _rows_to_dicts(rows),
            "pagination": {
                "page":        page,
                "page_size":   page_size,
                "total":       total,
                "total_pages": total_pages,
                "has_next":    page < total_pages,
                "has_prev":    page > 1,
            },
        })
    finally:
        conn.close()


# ─── Travel Endpoints ──────────────────────────────────────────────────────────

@app.route("/api/travel/persons")
def travel_persons():
    if not os.path.exists(TRAVEL_DB):
        return jsonify({"data": [], "pagination": {}}), 200
    page, page_size = _page_params()
    nationality = request.args.get("nationality")
    where  = "nationality = ?" if nationality else ""
    params = (nationality,)    if nationality else ()
    conn   = get_connection(TRAVEL_DB)
    try:
        return _paginated_response(conn, "persons", page, page_size, where, "id DESC", params)
    finally:
        conn.close()


@app.route("/api/travel/families")
def travel_families():
    if not os.path.exists(TRAVEL_DB):
        return jsonify({"data": [], "pagination": {}}), 200
    page, page_size = _page_params()
    conn = get_connection(TRAVEL_DB)
    try:
        return _paginated_response(conn, "families", page, page_size)
    finally:
        conn.close()


@app.route("/api/travel/documents")
def travel_documents():
    if not os.path.exists(TRAVEL_DB):
        return jsonify({"data": [], "pagination": {}}), 200
    page, page_size = _page_params()
    doc_type = request.args.get("doc_type")
    where    = "doc_type = ?" if doc_type else ""
    params   = (doc_type,)    if doc_type else ()
    conn     = get_connection(TRAVEL_DB)
    try:
        return _paginated_response(conn, "documents_travel", page, page_size, where, "id DESC", params)
    finally:
        conn.close()


# ─── HTML UI ──────────────────────────────────────────────────────────────────
# Server-rendered pages with DaisyUI (CDN) + HTMX (CDN). No build step.
# Lives alongside the JSON API on the same port (default 7845).

# Dashboard stats + action panel queries live in core.business.stats.
# Local thin wrappers preserve the existing leading-underscore signatures so
# routes don't have to change.
from core.business.stats import (
    logistics_stats as _ls_impl,
    logistics_recent_activity as _lra_impl,
    logistics_action_panel as _lap_impl,
    logistics_recent_containers as _lrc_impl,
)

def _logistics_stats() -> dict:
    return _ls_impl(LOGISTICS_DB)

def _logistics_recent_activity(limit: int = 15) -> list[dict]:
    return _lra_impl(LOGISTICS_DB, limit)

def _logistics_action_panel() -> dict:
    return _lap_impl(LOGISTICS_DB)

def _logistics_recent_containers(limit: int = 25) -> list[dict]:
    return _lrc_impl(LOGISTICS_DB, limit)


@app.route("/")
def ui_mode_picker():
    """Splash — pick logistics or travel."""
    return render_template("mode_picker.html")


@app.route("/logistics")
def ui_logistics_dashboard():
    """Logistics dashboard — operator action panel + recent containers."""
    return render_template(
        "logistics/dashboard.html",
        mode="logistics",
        stats=_logistics_stats(),
        action=_logistics_action_panel(),
        containers=_logistics_recent_containers(25),
        activity=_logistics_recent_activity(15),
    )


# ─── Travel mode (P4) ─────────────────────────────────────────────────────────
from core.business.stats import (
    travel_stats as _ts_impl,
    travel_recent_persons as _trp_impl,
)

def _travel_stats() -> dict:
    return _ts_impl(TRAVEL_DB)

def _travel_recent_persons(limit: int = 25) -> list[dict]:
    return _trp_impl(TRAVEL_DB, limit)


@app.route("/travel")
def ui_travel_dashboard():
    return render_template(
        "travel/dashboard.html",
        mode="travel",
        stats=_travel_stats(),
        persons=_travel_recent_persons(20),
        db_present=os.path.exists(TRAVEL_DB),
    )


@app.route("/travel/upload", methods=["GET"])
def ui_travel_upload():
    return render_template("travel/upload.html", mode="travel")


@app.route("/travel/upload", methods=["POST"])
def ui_travel_upload_post():
    files = request.files.getlist("files")
    doc_type = request.form.get("doc_type", "UNKNOWN")
    if not files:
        return render_template("logistics/_upload_error.html",
                               error="No files received"), 400

    travel_input = os.environ.get(
        "BRUNS_TRAVEL_INPUT_DIR",
        os.path.join(ROOT_DIR, "data", "input", "travel"),
    )
    os.makedirs(travel_input, exist_ok=True)

    submitted = []
    for f in files:
        if not f or not f.filename:
            continue
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ALLOWED_EXT:
            continue
        safe_name = secure_filename(f.filename) or f"upload{ext}"
        unique_name = f"{secrets.token_hex(4)}_{safe_name}"
        save_path = os.path.join(travel_input, unique_name)
        f.save(save_path)
        job_id = job_tracker.submit_job(save_path, module="travel", doc_type=doc_type)
        submitted.append({"job_id": job_id, "filename": safe_name})

    if not submitted:
        return render_template("logistics/_upload_error.html",
                               error="No accepted files (allowed: .pdf, .png, .jpg, .jpeg, .webp)"), 400

    # Use a travel-specific progress fragment so polling hits /travel/process-status/
    return render_template("travel/_upload_progress.html", jobs=submitted)


@app.route("/travel/process-status/<job_id>")
def ui_travel_process_status(job_id: str):
    job = job_tracker.get_job(job_id)
    if not job:
        return f'<div class="alert alert-error">Job {job_id} not found</div>', 404
    return render_template(
        "travel/_progress_row.html",
        job=job,
        percent=job_tracker.progress_percent(job),
    )


@app.route("/travel/persons")
def ui_travel_persons():
    if not os.path.exists(TRAVEL_DB):
        return render_template("travel/persons.html", mode="travel",
                               persons=[], q="", total=0, db_present=False)
    q = (request.args.get("q") or "").strip()
    conn = get_connection(TRAVEL_DB)
    try:
        if q:
            rows = conn.execute(
                """SELECT p.id, p.full_name, p.dob, p.nationality, p.gender,
                          f.family_name,
                          (SELECT COUNT(*) FROM documents_travel WHERE person_id=p.id) AS n_docs
                   FROM persons p LEFT JOIN families f ON f.id=p.family_id
                   WHERE p.full_name LIKE ? OR p.nationality LIKE ?
                      OR f.family_name LIKE ?
                   ORDER BY p.id DESC LIMIT 200""",
                (f"%{q}%", f"%{q}%", f"%{q}%"),
            ).fetchall()
        else:
            rows = conn.execute(
                """SELECT p.id, p.full_name, p.dob, p.nationality, p.gender,
                          f.family_name,
                          (SELECT COUNT(*) FROM documents_travel WHERE person_id=p.id) AS n_docs
                   FROM persons p LEFT JOIN families f ON f.id=p.family_id
                   ORDER BY p.id DESC LIMIT 200"""
            ).fetchall()
        total = conn.execute("SELECT COUNT(*) FROM persons").fetchone()[0]
        return render_template("travel/persons.html", mode="travel",
                               persons=[dict(r) for r in rows], q=q,
                               total=total, db_present=True)
    finally:
        conn.close()


@app.route("/travel/families")
def ui_travel_families():
    if not os.path.exists(TRAVEL_DB):
        return render_template("travel/families.html", mode="travel",
                               families=[], db_present=False)
    conn = get_connection(TRAVEL_DB)
    try:
        rows = conn.execute(
            """SELECT f.id, f.family_name, f.case_reference, f.address, f.notes,
                      f.case_status, f.next_action, f.next_action_date,
                      (SELECT COUNT(*) FROM persons WHERE family_id=f.id) AS n_persons,
                      (SELECT COUNT(*) FROM documents_travel WHERE family_id=f.id) AS n_docs
               FROM families f
               ORDER BY f.id DESC"""
        ).fetchall()
        return render_template("travel/families.html", mode="travel",
                               families=[dict(r) for r in rows], db_present=True)
    finally:
        conn.close()


# ─── P3-A: Expiry heatmap calendar (next 12 months) ─────────────────────────

@app.route("/travel/calendar")
def ui_travel_calendar():
    """Travel document expiry heatmap — see at-a-glance which months are heavy."""
    if not os.path.exists(TRAVEL_DB):
        return render_template("travel/calendar.html", mode="travel",
                               months=[], db_present=False)
    from datetime import date, timedelta
    today = date.today()
    months = []
    for i in range(12):
        y = today.year + ((today.month - 1 + i) // 12)
        m = ((today.month - 1 + i) % 12) + 1
        months.append({
            "year": y, "month": m, "key": f"{y:04d}-{m:02d}",
            "label": ["Jan","Feb","Mar","Apr","May","Jun",
                      "Jul","Aug","Sep","Oct","Nov","Dec"][m-1],
            "docs": [], "count": 0,
        })

    conn = get_connection(TRAVEL_DB)
    try:
        rows = conn.execute(
            """SELECT dt.id, dt.doc_type, dt.doc_number, dt.expiry_date,
                      p.full_name, p.id AS person_id, f.family_name
               FROM documents_travel dt
               LEFT JOIN persons p ON p.id = dt.person_id
               LEFT JOIN families f ON f.id = dt.family_id
               WHERE dt.expiry_date IS NOT NULL AND dt.expiry_date != ''"""
        ).fetchall()
    finally:
        conn.close()

    by_key = {m["key"]: m for m in months}
    expired = []
    for r in rows:
        try:
            exp = r["expiry_date"][:10]
            y, mo = int(exp[:4]), int(exp[5:7])
        except Exception:
            continue
        days_until = (date(y, mo, int(exp[8:10] or "1")) - today).days if len(exp) >= 10 else None
        item = {
            "id": r["id"], "doc_type": r["doc_type"], "doc_number": r["doc_number"],
            "expiry_date": exp, "person": r["full_name"], "person_id": r["person_id"],
            "family": r["family_name"], "days_until": days_until,
        }
        if days_until is not None and days_until < 0:
            expired.append(item)
            continue
        key = f"{y:04d}-{mo:02d}"
        if key in by_key:
            by_key[key]["docs"].append(item)
            by_key[key]["count"] += 1

    max_count = max((m["count"] for m in months), default=0)
    for m in months:
        m["docs"].sort(key=lambda d: d["expiry_date"])
        if max_count == 0:
            m["heat"] = 0
        else:
            ratio = m["count"] / max_count
            m["heat"] = 4 if ratio >= 0.75 else 3 if ratio >= 0.5 else 2 if ratio >= 0.25 else 1 if ratio > 0 else 0

    return render_template("travel/calendar.html", mode="travel",
                           months=months, expired=expired, db_present=True)


# ─── P3-B/C/E: Family case detail (helpers in core.business.completeness) ──
# CASE_STATUS_FLOW is referenced by the family detail + update routes;
# REQUIRED_DOCS_BY_ROLE is consumed inside family_completeness() so it
# doesn't need to be re-exported here.
from core.business.completeness import (
    CASE_STATUS_FLOW,
    family_completeness as _family_completeness,
)


@app.route("/travel/families/<int:family_id>", methods=["GET"])
def ui_travel_family_detail(family_id: int):
    if not os.path.exists(TRAVEL_DB):
        abort(404)
    conn = get_connection(TRAVEL_DB)
    try:
        fam = conn.execute("SELECT * FROM families WHERE id = ?", (family_id,)).fetchone()
        if not fam:
            abort(404, f"Family {family_id} not found")
        completeness = _family_completeness(conn, family_id)
        # Recent documents in this case
        recent_docs = conn.execute(
            """SELECT dt.id, dt.doc_type, dt.doc_number, dt.expiry_date,
                      p.full_name, p.id AS person_id
               FROM documents_travel dt
               LEFT JOIN persons p ON p.id = dt.person_id
               WHERE dt.family_id = ?
               ORDER BY dt.id DESC LIMIT 50""",
            (family_id,),
        ).fetchall()
    finally:
        conn.close()
    fam_dict = dict(fam)
    return render_template("travel/family_detail.html", mode="travel",
                           f=fam_dict, completeness=completeness,
                           recent_docs=[dict(r) for r in recent_docs],
                           status_flow=CASE_STATUS_FLOW)


@app.route("/travel/families/<int:family_id>", methods=["POST"])
def ui_travel_family_update(family_id: int):
    """Update case_status, next_action, next_action_date with completeness gate."""
    if not os.path.exists(TRAVEL_DB):
        abort(404)
    new_status = (request.form.get("case_status") or "").strip().upper()
    next_action = (request.form.get("next_action") or "").strip()
    next_action_date = (request.form.get("next_action_date") or "").strip()

    conn = get_connection(TRAVEL_DB)
    try:
        completeness = _family_completeness(conn, family_id)
        # Gate: cannot advance to IN_REVIEW or beyond if not 100% complete
        gated = ["IN_REVIEW", "SUBMITTED", "APPROVED"]
        if new_status in gated and not completeness["can_advance"]:
            return render_template(
                "travel/_save_result.html", ok=False,
                msg=f"Cannot advance to {new_status} — {len(completeness['blockers'])} blocker(s): "
                    + "; ".join(completeness["blockers"][:3]),
            )
        if new_status and new_status not in CASE_STATUS_FLOW:
            new_status = "COLLECTING"

        sets, params = [], []
        if new_status:
            sets.append("case_status = ?")
            params.append(new_status)
        sets.append("next_action = ?")
        params.append(next_action or None)
        sets.append("next_action_date = ?")
        params.append(next_action_date or None)
        params.append(family_id)
        conn.execute(f"UPDATE families SET {', '.join(sets)} WHERE id = ?", params)
        try:
            conn.execute(
                """INSERT INTO audit_log (action, actor, entity_type, entity_id, after_json, timestamp)
                   VALUES ('family_update', 'operator', 'families', ?, ?, datetime('now'))""",
                (str(family_id), json.dumps({"case_status": new_status,
                                             "next_action": next_action,
                                             "next_action_date": next_action_date})),
            )
        except Exception:
            pass
        conn.commit()
        return render_template("travel/_save_result.html", ok=True,
                               msg=f"✓ Saved — status: {new_status or 'unchanged'}")
    except Exception as e:
        return render_template("travel/_save_result.html", ok=False,
                               msg=f"Save failed: {e}")
    finally:
        conn.close()


# ─── P3-D: Family dossier ZIP export (single family) ────────────────────────

@app.route("/travel/families/<int:family_id>/export")
def ui_travel_family_export(family_id: int):
    """Build a fresh ZIP containing all source files for one family + an XLSX summary."""
    if not os.path.exists(TRAVEL_DB):
        abort(404)
    import io as _io
    import zipfile

    conn = get_connection(TRAVEL_DB)
    try:
        fam = conn.execute("SELECT * FROM families WHERE id = ?", (family_id,)).fetchone()
        if not fam:
            abort(404)
        persons = conn.execute(
            "SELECT * FROM persons WHERE family_id = ? ORDER BY id",
            (family_id,),
        ).fetchall()
        # Bridge from documents_travel → documents (source_file)
        rows = conn.execute(
            """SELECT dt.id, dt.doc_type, dt.doc_number, dt.expiry_date,
                      dt.person_id, p.full_name,
                      d.source_file
               FROM documents_travel dt
               LEFT JOIN persons p ON p.id = dt.person_id
               LEFT JOIN documents d ON d.id = dt.original_doc_id
               WHERE dt.family_id = ?""",
            (family_id,),
        ).fetchall()
    finally:
        conn.close()

    fam_name = (fam["family_name"] or f"family_{family_id}").replace("/", "_")
    buf = _io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Source files organised by member
        for r in rows:
            src = r["source_file"]
            if src and os.path.isfile(src):
                member = (r["full_name"] or f"person_{r['person_id'] or 0}").replace("/", "_")
                arc = f"{fam_name}/{member}/{r['doc_type'] or 'DOC'}_{os.path.basename(src)}"
                try:
                    zf.write(src, arc)
                except Exception:
                    pass
        # Summary CSV (no pandas dependency for this hot path)
        import csv
        sb = _io.StringIO()
        w = csv.writer(sb)
        w.writerow(["person_id", "full_name", "doc_type", "doc_number", "expiry_date"])
        for r in rows:
            w.writerow([r["person_id"], r["full_name"], r["doc_type"],
                        r["doc_number"], r["expiry_date"]])
        zf.writestr(f"{fam_name}/_summary.csv", sb.getvalue())
        # Family info
        zf.writestr(f"{fam_name}/_family.txt",
                    f"Family: {fam['family_name']}\n"
                    f"Case ref: {fam['case_reference']}\n"
                    f"Address: {fam['address']}\n"
                    f"Status: {fam['case_status']}\n"
                    f"Next action: {fam['next_action']} ({fam['next_action_date']})\n"
                    f"Members: {len(persons)}\n"
                    f"Documents: {len(rows)}\n")
    buf.seek(0)
    from datetime import datetime as _dt
    fname = f"{fam_name}_{_dt.now().strftime('%Y%m%d_%H%M')}.zip"
    return send_file(buf, mimetype="application/zip",
                     as_attachment=True, download_name=fname)


# ─── File serve + detail pages (split view + live edit) ──────────────────────
import mimetypes
from flask import send_file


def _doc_db_for_module(module: str) -> str:
    return TRAVEL_DB if module == "travel" else LOGISTICS_DB


@app.route("/files/<module>/<int:doc_id>")
def ui_serve_original_file(module: str, doc_id: int):
    """Serve the original uploaded file referenced by documents.id.

    Security: only files referenced by an existing row are served — no
    arbitrary path access.
    """
    if module not in ("travel", "logistics"):
        abort(404)
    db = _doc_db_for_module(module)
    if not os.path.exists(db):
        abort(404)
    conn = get_connection(db)
    try:
        row = conn.execute(
            "SELECT source_file FROM documents WHERE id = ? AND module = ?",
            (doc_id, module),
        ).fetchone()
    finally:
        conn.close()
    if not row or not row["source_file"]:
        abort(404)
    path = row["source_file"]
    if not os.path.isfile(path):
        abort(404, f"Source file moved or deleted: {os.path.basename(path)}")
    mime, _ = mimetypes.guess_type(path)
    return send_file(path, mimetype=mime or "application/octet-stream")


# ─── Pure business helpers (TD8 phase 1 extracts) ──────────────────────────
# Routes use these under their leading-underscore aliases; the heavy lifting
# now lives in core.business.* and is unit-testable without Flask.
from core.business.demurrage import (
    demurrage_info as _demurrage_info,
    free_days_from_documents as _free_days_from_documents,
)
from core.business.reconcile import (
    compute_diff as _compute_diff,
    flatten_diff as _flatten_diff,
    reconcile_siblings as _reconcile_siblings_impl,
)


# ─── P2-B: Natural Language Query — "Ask your data" ─────────────────────────
# Schema description, SELECT-only safety gate, and LLM invocation all live
# in core.business.nlsql. This route stays thin: validate, execute under
# `PRAGMA query_only = ON`, render the results fragment.
from core.business.nlsql import (
    generate_sql_from_question as _generate_sql_from_question,
    validate_select as _nl_validate_select,
)


@app.route("/logistics/ask", methods=["POST"])
def ui_logistics_ask():
    """P2-B: NL2SQL endpoint. Accepts a question, returns a results fragment."""
    question = (request.form.get("q") or "").strip()
    if not question:
        return render_template("logistics/_ask_result.html",
                               error="Empty question.", rows=[], cols=[], sql="")

    # 1. LLM call
    try:
        sql = _generate_sql_from_question(question)
    except Exception as e:
        return render_template("logistics/_ask_result.html",
                               error=f"LLM not reachable: {e}",
                               rows=[], cols=[], sql="", question=question)

    # 2. Safety gate (TD7 — see core.business.nlsql.validate_select)
    ok, err, sql_clean = _nl_validate_select(sql)
    if not ok:
        return render_template("logistics/_ask_result.html",
                               error=err, rows=[], cols=[],
                               sql=sql, question=question)

    # 3. Execute under PRAGMA query_only as defense in depth.
    if not os.path.exists(LOGISTICS_DB):
        return render_template("logistics/_ask_result.html",
                               error="Logistics database not found.", rows=[], cols=[], sql=sql)
    conn = None
    try:
        conn = get_connection(LOGISTICS_DB)
        conn.execute("PRAGMA query_only = ON")
        cursor = conn.execute(sql_clean)
        cols = [d[0] for d in (cursor.description or [])]
        rows = [list(r) for r in (cursor.fetchmany(50) or [])]
    except Exception as e:
        return render_template("logistics/_ask_result.html",
                               error=f"SQL error: {e}", rows=[], cols=[],
                               sql=sql, question=question)
    finally:
        if conn is not None:
            try: conn.close()
            except Exception: pass

    return render_template("logistics/_ask_result.html",
                           error=None, rows=rows, cols=cols,
                           sql=sql, question=question)


# ─── P2-A: Bounding box field visualization ──────────────────────────────────
# Color palette + path → color mapping live in core.business.bbox.
from core.business.bbox import field_color as _field_color  # noqa: F401


@app.route("/files/<module>/<int:doc_id>/annotated")
def ui_serve_annotated(module: str, doc_id: int):
    """Render the source PDF with highlight overlays on all extracted field values.

    Query params:
      page (int, default 0) — which PDF page to render
      field (str, optional) — if given, only highlight this specific field path

    Returns a PNG image of the annotated page.
    Uses PyMuPDF's built-in text search + highlight annotation.
    """
    if module not in ("travel", "logistics"):
        abort(404)
    db = _doc_db_for_module(module)
    if not os.path.exists(db):
        abort(404)

    conn = get_connection(db)
    try:
        row = conn.execute(
            "SELECT source_file, extracted_json FROM documents WHERE id = ? AND module = ?",
            (doc_id, module),
        ).fetchone()
    finally:
        conn.close()

    if not row or not row["source_file"]:
        abort(404)
    path = row["source_file"]
    if not os.path.isfile(path):
        abort(404, f"Source file missing: {os.path.basename(path)}")

    try:
        extracted = json.loads(row["extracted_json"] or "{}")
    except Exception:
        extracted = {}

    try:
        import fitz
    except ImportError:
        abort(500, "PyMuPDF not installed")

    page_num = max(0, int(request.args.get("page", 0)))
    active_field = request.args.get("field", None)

    try:
        doc = fitz.open(path)
    except Exception as e:
        abort(500, f"Cannot open PDF: {e}")

    if page_num >= len(doc):
        page_num = 0
    page = doc[page_num]

    # Collect (value, color) pairs to highlight
    flat = _flatten_diff(extracted)
    highlights = []

    for field_path, value in flat.items():
        if value is None or value == "" or isinstance(value, bool):
            continue
        # Only highlight the active field if one is specified
        if active_field and field_path != active_field:
            continue
        val_str = str(value).strip()
        if len(val_str) < 3:
            continue
        color = _field_color(field_path)
        # Boost opacity if this is the active field
        alpha = 0.7 if (active_field and field_path == active_field) else 0.35
        highlights.append((val_str, color, alpha))

    # Apply highlights: search for each value, add annotation
    for val_str, color, alpha in highlights:
        try:
            instances = page.search_for(val_str, quads=True)
            for inst in instances:
                ann = page.add_highlight_annot(inst)
                ann.set_colors(stroke=color)
                ann.set_opacity(alpha)
                ann.update()
        except Exception:
            continue

    # Render annotated page to PNG
    mat = fitz.Matrix(2.0, 2.0)  # 144 DPI — crisp on retina, reasonable file size
    pix = page.get_pixmap(matrix=mat, alpha=False)
    png_bytes = pix.tobytes("png")
    doc.close()

    from flask import Response as FlaskResponse
    return FlaskResponse(png_bytes, mimetype="image/png",
                         headers={"Cache-Control": "no-store"})


def _fetch_travel_doc_detail(travel_doc_id: int) -> dict | None:
    """Pull a documents_travel row + its linked documents row + person."""
    if not os.path.exists(TRAVEL_DB):
        return None
    conn = get_connection(TRAVEL_DB)
    try:
        row = conn.execute(
            """
            SELECT
                dt.id              AS travel_id,
                dt.original_doc_id AS doc_id,
                dt.doc_type        AS doc_type,
                dt.doc_number      AS doc_number,
                dt.expiry_date     AS expiry_date,
                dt.mrz_raw         AS mrz_raw,
                dt.person_id       AS person_id,
                dt.family_id       AS family_id,
                p.full_name        AS person_name,
                p.dob              AS person_dob,
                p.nationality      AS person_nationality,
                p.gender           AS person_gender,
                d.source_file      AS source_file,
                d.extracted_json   AS extracted_json,
                d.confidence       AS confidence,
                d.created_at       AS created_at
            FROM documents_travel dt
            LEFT JOIN persons   p ON p.id = dt.person_id
            LEFT JOIN documents d ON d.id = dt.original_doc_id
            WHERE dt.id = ?
            """,
            (travel_doc_id,),
        ).fetchone()
        return dict(row) if row else None
    finally:
        conn.close()


@app.route("/travel/documents/<int:travel_doc_id>", methods=["GET"])
def ui_travel_document_detail(travel_doc_id: int):
    detail = _fetch_travel_doc_detail(travel_doc_id)
    if not detail:
        abort(404, f"Travel document {travel_doc_id} not found")
    parsed_extracted: dict = {}
    if detail.get("extracted_json"):
        try:
            parsed_extracted = json.loads(detail["extracted_json"])
        except Exception:
            parsed_extracted = {}
    return render_template(
        "travel/document_detail.html",
        mode="travel",
        d=detail,
        extracted=parsed_extracted,
        ext=os.path.splitext(detail.get("source_file") or "")[1].lower(),
        basename=os.path.basename(detail.get("source_file") or ""),
    )


@app.route("/travel/documents/<int:travel_doc_id>", methods=["POST"])
def ui_travel_document_update(travel_doc_id: int):
    """Inline-edit handler. Updates documents_travel + the JSON in documents."""
    if not os.path.exists(TRAVEL_DB):
        abort(404)
    fields = {
        "doc_type":    (request.form.get("doc_type") or "").strip().upper() or None,
        "doc_number":  (request.form.get("doc_number") or "").strip() or None,
        "expiry_date": (request.form.get("expiry_date") or "").strip() or None,
        "mrz_raw":     (request.form.get("mrz_raw") or "").strip() or None,
    }
    conn = get_connection(TRAVEL_DB)
    try:
        # Update domain row
        conn.execute(
            """UPDATE documents_travel
                  SET doc_type    = ?,
                      doc_number  = ?,
                      expiry_date = ?,
                      mrz_raw     = ?
                WHERE id = ?""",
            (fields["doc_type"], fields["doc_number"],
             fields["expiry_date"], fields["mrz_raw"], travel_doc_id),
        )
        # Mirror into the source documents.extracted_json so re-projecting
        # (or BI export) sees the corrected values too.
        link = conn.execute(
            "SELECT original_doc_id FROM documents_travel WHERE id = ?",
            (travel_doc_id,),
        ).fetchone()
        if link and link["original_doc_id"]:
            existing = conn.execute(
                "SELECT extracted_json FROM documents WHERE id = ?",
                (link["original_doc_id"],),
            ).fetchone()
            data = {}
            if existing and existing["extracted_json"]:
                try:
                    data = json.loads(existing["extracted_json"])
                except Exception:
                    data = {}
            for k, v in fields.items():
                if v is not None:
                    if k == "doc_number":
                        data["document_number"] = v
                    elif k == "doc_type":
                        data["document_type"] = v
                    elif k == "expiry_date":
                        data["expiry_date"] = v
                    elif k == "mrz_raw":
                        data["mrz_raw"] = v
            conn.execute(
                "UPDATE documents SET extracted_json = ? WHERE id = ?",
                (json.dumps(data), link["original_doc_id"]),
            )
        conn.commit()
        return render_template("travel/_save_result.html", ok=True,
                               msg=f"Saved — {sum(1 for v in fields.values() if v is not None)} field(s) updated.")
    except Exception as e:
        return render_template("travel/_save_result.html", ok=False,
                               msg=f"Save failed: {type(e).__name__}: {e}")
    finally:
        conn.close()


@app.route("/travel/persons/<int:person_id>", methods=["GET"])
def ui_travel_person_detail(person_id: int):
    if not os.path.exists(TRAVEL_DB):
        abort(404)
    conn = get_connection(TRAVEL_DB)
    try:
        person = conn.execute(
            """SELECT p.id, p.full_name, p.dob, p.nationality, p.gender,
                      p.family_id, f.family_name
               FROM persons p
               LEFT JOIN families f ON f.id = p.family_id
               WHERE p.id = ?""",
            (person_id,),
        ).fetchone()
        if not person:
            abort(404, f"Person {person_id} not found")
        docs = conn.execute(
            """SELECT id, doc_type, doc_number, expiry_date, original_doc_id
               FROM documents_travel
               WHERE person_id = ?
               ORDER BY id DESC""",
            (person_id,),
        ).fetchall()
        return render_template(
            "travel/person_detail.html",
            mode="travel",
            person=dict(person),
            docs=[dict(d) for d in docs],
        )
    finally:
        conn.close()


@app.route("/travel/persons/<int:person_id>", methods=["POST"])
def ui_travel_person_update(person_id: int):
    if not os.path.exists(TRAVEL_DB):
        abort(404)
    fields = {
        "full_name":   (request.form.get("full_name") or "").strip() or None,
        "dob":         (request.form.get("dob") or "").strip() or None,
        "nationality": (request.form.get("nationality") or "").strip() or None,
        "gender":      (request.form.get("gender") or "").strip() or None,
    }
    normalized = (fields["full_name"] or "").lower().strip() or None
    conn = get_connection(TRAVEL_DB)
    try:
        conn.execute(
            """UPDATE persons
                  SET full_name = ?, normalized_name = ?,
                      dob = ?, nationality = ?, gender = ?
                WHERE id = ?""",
            (fields["full_name"], normalized,
             fields["dob"], fields["nationality"], fields["gender"], person_id),
        )
        conn.commit()
        return render_template("travel/_save_result.html", ok=True,
                               msg=f"Person saved — {sum(1 for v in fields.values() if v) } field(s).")
    except Exception as e:
        return render_template("travel/_save_result.html", ok=False,
                               msg=f"Save failed: {type(e).__name__}: {e}")
    finally:
        conn.close()


@app.route("/travel/documents")
def ui_travel_documents():
    if not os.path.exists(TRAVEL_DB):
        return render_template("travel/documents.html", mode="travel",
                               documents=[], doc_type_filter="", db_present=False)
    dt = (request.args.get("type") or "").strip()
    conn = get_connection(TRAVEL_DB)
    try:
        if dt:
            rows = conn.execute(
                """SELECT d.id, d.doc_type, d.doc_number, d.expiry_date,
                          p.full_name, p.nationality, f.family_name
                   FROM documents_travel d
                   LEFT JOIN persons p ON p.id = d.person_id
                   LEFT JOIN families f ON f.id = d.family_id
                   WHERE d.doc_type = ?
                   ORDER BY d.id DESC LIMIT 300""", (dt,),
            ).fetchall()
        else:
            rows = conn.execute(
                """SELECT d.id, d.doc_type, d.doc_number, d.expiry_date,
                          p.full_name, p.nationality, f.family_name
                   FROM documents_travel d
                   LEFT JOIN persons p ON p.id = d.person_id
                   LEFT JOIN families f ON f.id = d.family_id
                   ORDER BY d.id DESC LIMIT 300"""
            ).fetchall()
        return render_template("travel/documents.html", mode="travel",
                               documents=[dict(r) for r in rows],
                               doc_type_filter=dt, db_present=True)
    finally:
        conn.close()


# ─── Settings (P3.4) ──────────────────────────────────────────────────────────
def _settings_overview() -> dict:
    """Collect read-only environment + config info for the settings page."""
    cfg_path = os.path.join(ROOT_DIR, "data", ".llm_config.json")
    cfg_present = os.path.exists(cfg_path)
    return {
        "data_dir":         DATA_DIR,
        "logistics_db":     LOGISTICS_DB,
        "logistics_db_ok":  os.path.exists(LOGISTICS_DB),
        "logistics_db_size": os.path.getsize(LOGISTICS_DB) if os.path.exists(LOGISTICS_DB) else 0,
        "travel_db":        TRAVEL_DB,
        "travel_db_ok":     os.path.exists(TRAVEL_DB),
        "travel_db_size":   os.path.getsize(TRAVEL_DB) if os.path.exists(TRAVEL_DB) else 0,
        "input_dir":        INPUT_DIR,
        "input_dir_ok":     os.path.isdir(INPUT_DIR),
        "input_count":      len([f for f in os.listdir(INPUT_DIR)
                                 if os.path.isfile(os.path.join(INPUT_DIR, f))])
                            if os.path.isdir(INPUT_DIR) else 0,
        "llm_cfg_path":     cfg_path,
        "llm_cfg_present":  cfg_present,
        "llm_cfg":          llm_config.load_config() if cfg_present else llm_config.DEFAULT_CONFIG,
        "active_jobs":      len([j for j in job_tracker.list_recent(50)
                                 if j.status in ("PENDING", "PROCESSING")]),
        "recent_jobs":      job_tracker.list_recent(10),
    }


@app.route("/logistics/settings", methods=["GET"])
def ui_logistics_settings():
    return render_template(
        "logistics/settings.html",
        mode="logistics",
        s=_settings_overview(),
    )


@app.route("/logistics/settings/clear-input", methods=["POST"])
def ui_logistics_clear_input():
    """Delete files in the input directory (already-processed PDFs).
    Does NOT delete from the database."""
    if not os.path.isdir(INPUT_DIR):
        return render_template("logistics/_settings_result.html",
                               ok=False, msg="Input directory not found")
    deleted = 0
    for fn in os.listdir(INPUT_DIR):
        p = os.path.join(INPUT_DIR, fn)
        if os.path.isfile(p):
            try:
                os.remove(p); deleted += 1
            except Exception:
                pass
    return render_template("logistics/_settings_result.html",
                           ok=True, msg=f"Deleted {deleted} input file(s)")


@app.route("/logistics/settings/purge-jobs", methods=["POST"])
def ui_logistics_purge_jobs():
    job_tracker._JOBS.clear()
    return render_template("logistics/_settings_result.html",
                           ok=True, msg="Cleared in-memory job history")


# ─── LLM config modal (P3.1) ──────────────────────────────────────────────────
from core.api import llm_config


@app.route("/llm/config", methods=["GET"])
def ui_llm_config_form():
    """Returns the modal body. HTMX swaps it into #llm_modal_body."""
    cfg = llm_config.load_config()
    return render_template(
        "llm/_modal_form.html",
        cfg=cfg,
        providers=llm_config.PROVIDERS,
        models=[],
        status=None,
    )


@app.route("/llm/test", methods=["POST"])
def ui_llm_test_connection():
    cfg = _llm_cfg_from_form(request.form)
    ok, msg, models = llm_config.test_connection(cfg)
    return render_template(
        "llm/_modal_form.html",
        cfg=cfg,
        providers=llm_config.PROVIDERS,
        models=models,
        status={"ok": ok, "msg": msg},
    )


@app.route("/llm/save", methods=["POST"])
def ui_llm_save():
    cfg = _llm_cfg_from_form(request.form)
    try:
        llm_config.save_config(cfg)
        return render_template(
            "llm/_modal_form.html",
            cfg=cfg,
            providers=llm_config.PROVIDERS,
            models=[],
            status={"ok": True, "msg": "Saved to data/.llm_config.json"},
        )
    except Exception as e:
        return render_template(
            "llm/_modal_form.html",
            cfg=cfg,
            providers=llm_config.PROVIDERS,
            models=[],
            status={"ok": False, "msg": f"Save failed: {e}"},
        )


def _llm_cfg_from_form(form) -> dict:
    """Coerce a multipart/form-encoded request into the cfg dict shape."""
    return {
        "provider":    form.get("provider", "ollama"),
        "base_url":    form.get("base_url", "http://localhost").strip(),
        "port":        int(form.get("port", 11434) or 11434),
        "model":       form.get("model", "").strip(),
        "api_key":     form.get("api_key", "").strip(),
        "temperature": float(form.get("temperature", 0.1) or 0.1),
        "timeout":     int(form.get("timeout", 120) or 120),
    }


# ─── Upload + HTMX progress polling (P2) ──────────────────────────────────────
import secrets
from werkzeug.utils import secure_filename
from core.api import job_tracker

ALLOWED_EXT = {".pdf", ".png", ".jpg", ".jpeg", ".webp"}
INPUT_DIR = os.environ.get(
    "BRUNS_INPUT_DIR",
    os.path.join(ROOT_DIR, "data", "input"),
)


@app.route("/logistics/upload", methods=["GET"])
def ui_logistics_upload():
    return render_template("logistics/upload.html", mode="logistics")


@app.route("/logistics/upload", methods=["POST"])
def ui_logistics_upload_post():
    """HTMX-aware upload handler. Spawns a job per file; returns the polling fragment."""
    files = request.files.getlist("files")
    if not files:
        return render_template("logistics/_upload_error.html",
                               error="No files received"), 400

    os.makedirs(INPUT_DIR, exist_ok=True)
    submitted = []

    for f in files:
        if not f or not f.filename:
            continue
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ALLOWED_EXT:
            continue
        # Prefix with a short random hash so concurrent uploads of identical
        # filenames don't clobber each other on disk.
        safe_name = secure_filename(f.filename) or f"upload{ext}"
        unique_name = f"{secrets.token_hex(4)}_{safe_name}"
        save_path = os.path.join(INPUT_DIR, unique_name)
        f.save(save_path)
        job_id = job_tracker.submit_job(save_path, module="logistics")
        submitted.append({"job_id": job_id, "filename": safe_name})

    if not submitted:
        return render_template("logistics/_upload_error.html",
                               error="No accepted files (allowed: .pdf, .png, .jpg, .jpeg, .webp)"), 400

    return render_template("logistics/_upload_progress.html", jobs=submitted)


@app.route("/logistics/process-status/<job_id>")
def ui_logistics_process_status(job_id: str):
    """HTMX polling target. Returns a single job's progress row."""
    job = job_tracker.get_job(job_id)
    if not job:
        return f'<div class="alert alert-error">Job {job_id} not found</div>', 404
    return render_template(
        "logistics/_progress_row.html",
        job=job,
        percent=job_tracker.progress_percent(job),
    )


# ─── Edit container (P3.2) ────────────────────────────────────────────────────
EDITABLE_FIELDS = {
    # Status & Delivery
    "statut_container":        "text",
    "date_livraison":          "date",
    "site_livraison":          "text",
    "livre_camion":            "text",
    "livre_chauffeur":         "text",
    # Transport
    "container_number":        "text",
    "size":                    "text",
    "seal_number":             "text",
    "date_depotement":         "date",
    "date_restitution":        "date",
    "restitue_camion":         "text",
    "restitue_chauffeur":      "text",
    "centre_restitution":      "text",
    # Demurrage & Billing
    "date_debut_surestarie":   "date",
    "date_restitution_estimative": "date",
    "nbr_jours_surestarie_estimes": "number",
    "nbr_jour_surestarie_facture":  "number",
    "montant_facture_check":   "number",
    "montant_facture_da":      "number",
    "taux_de_change":          "number",
    "n_facture_cm":            "text",
    # Douane
    "date_declaration_douane": "date",
    "date_liberation_douane":  "date",
    "nbr_jours_perdu_douane":  "number",
    # Notes
    "commentaire":             "textarea",
}


def _fetch_container_row(container_id: int) -> dict | None:
    if not os.path.exists(LOGISTICS_DB):
        return None
    conn = get_connection(LOGISTICS_DB)
    try:
        row = conn.execute(
            """
            SELECT c.*, s.tan, s.compagnie_maritime, s.port, s.vessel,
                   s.etd, s.eta, s.source_file
            FROM containers c JOIN shipments s ON s.id = c.shipment_id
            WHERE c.id = ?
            """,
            (container_id,),
        ).fetchone()
        return dict(row) if row else None
    finally:
        conn.close()


# ─── Logistics document detail (full extracted JSON, editable) ────────────────
# Flat-form helpers extracted to core.business.forms.
from core.business.forms import (
    flatten_for_form as _flatten_for_form,
    set_nested as _set_nested,
)


def _semantic_search_ids(q: str, module: str, n: int = 50) -> list[int]:
    """P2-G: ChromaDB semantic search → list of document IDs ranked by similarity.

    Returns [] if ChromaDB or the vector store isn't available — caller falls
    back to LIKE matching.
    """
    try:
        from core.search.vector_db import VectorSearchEngine
        vector_dir = os.path.join(DATA_DIR, "vector")
        if not os.path.isdir(vector_dir):
            return []
        engine = VectorSearchEngine(vector_dir)
        hits = engine.search(q, module=module, n_results=n)
        return [h["document_id"] for h in hits if h.get("document_id") is not None]
    except Exception as e:
        log.warning("semantic search fallback to LIKE matching: %s", e)
        return []


@app.route("/logistics/documents")
def ui_logistics_documents_list():
    """List ALL extracted logistics documents, with optional semantic search."""
    if not os.path.exists(LOGISTICS_DB):
        return render_template("logistics/documents.html", mode="logistics",
                               docs=[], q="", db_present=False, semantic=False)
    q = (request.args.get("q") or "").strip()
    semantic = request.args.get("semantic") == "1" and bool(q)

    semantic_ids = _semantic_search_ids(q, "logistics") if semantic else []
    semantic_used = bool(semantic_ids)

    conn = get_connection(LOGISTICS_DB)
    try:
        if semantic_used:
            # Pull only the docs returned by ChromaDB, preserving rank order
            placeholders = ",".join("?" * len(semantic_ids))
            rows = conn.execute(
                f"""SELECT id, type, source_file, confidence, created_at, extracted_json
                    FROM documents
                    WHERE module = 'logistics' AND id IN ({placeholders})""",
                semantic_ids,
            ).fetchall()
            row_by_id = {r["id"]: r for r in rows}
            rows = [row_by_id[i] for i in semantic_ids if i in row_by_id]
        else:
            rows = conn.execute(
                """SELECT id, type, source_file, confidence, created_at, extracted_json
                   FROM documents
                   WHERE module = 'logistics'
                     AND (extracted_json IS NULL OR length(extracted_json) > 0)
                   ORDER BY id DESC LIMIT 500"""
            ).fetchall()
    finally:
        conn.close()

    docs = []
    for r in rows:
        try:
            ed = json.loads(r["extracted_json"] or "{}")
        except Exception:
            ed = {}
        tan = ed.get("tan_number") or ""
        if q and not semantic_used:
            if q.lower() not in (tan or "").lower() \
               and q.lower() not in (r["type"] or "").lower() \
               and q.lower() not in os.path.basename(r["source_file"] or "").lower():
                continue
        docs.append({
            "id": r["id"],
            "type": r["type"],
            "tan": tan,
            "vessel": ed.get("vessel_name"),
            "carrier": ed.get("carrier"),
            "etd": ed.get("etd"),
            "eta": ed.get("eta"),
            "n_containers": len(ed.get("containers") or []),
            "source_file": r["source_file"],
            "basename": os.path.basename(r["source_file"] or ""),
            "confidence": r["confidence"],
            "created_at": r["created_at"],
        })
    return render_template("logistics/documents.html", mode="logistics",
                           docs=docs, q=q, db_present=True,
                           semantic=semantic, semantic_used=semantic_used)


# ─── P2-D: Cross-document reconciliation (extracted to core.business.reconcile) ─
def _reconcile_siblings(doc_id: int, tan: str, current: dict) -> dict:
    """Thin wrapper that injects LOGISTICS_DB so the caller signature is
    unchanged. Implementation lives in core.business.reconcile."""
    return _reconcile_siblings_impl(LOGISTICS_DB, doc_id, tan, current)


@app.route("/logistics/documents/<int:doc_id>", methods=["GET"])
def ui_logistics_document_detail(doc_id: int):
    """Split view: PDF on left, ALL extracted JSON fields editable on right."""
    if not os.path.exists(LOGISTICS_DB):
        abort(404)
    conn = get_connection(LOGISTICS_DB)
    try:
        row = conn.execute(
            "SELECT * FROM documents WHERE id = ? AND module = 'logistics'",
            (doc_id,),
        ).fetchone()
    finally:
        conn.close()
    if not row:
        abort(404, f"Logistics document {doc_id} not found")
    try:
        extracted = json.loads(row["extracted_json"] or "{}")
    except Exception:
        extracted = {}
    flat_fields = _flatten_for_form(extracted)
    extracted_json_pretty = json.dumps(extracted, indent=2, ensure_ascii=False, default=str)
    tan = extracted.get("tan_number")
    reconciliation = _reconcile_siblings(doc_id, tan, extracted)
    return render_template(
        "logistics/document_detail.html",
        mode="logistics",
        d=dict(row),
        extracted=extracted,
        flat_fields=flat_fields,
        extracted_json=extracted_json_pretty,
        basename=os.path.basename(row["source_file"] or ""),
        ext=os.path.splitext(row["source_file"] or "")[1].lower(),
        tier=confidence_tier(row["confidence"]),
        reconciliation=reconciliation,
    )


# ─── P1-B: Three-tier confidence routing ─────────────────────────────────────
def confidence_tier(score) -> dict:
    """Map a confidence score to a tier: auto | review | attention | unknown.

    Returns {tier, label, color, css_class} for UI rendering.
    Industry standard from Microsoft Azure DI / Parseur:
      ≥ 0.90 → auto-approved
      0.60–0.89 → review queue
      < 0.60 → needs attention
    """
    if score is None:
        return {"tier": "unknown", "label": "no score",
                "color": "ghost", "css_class": "badge-ghost"}
    try:
        s = float(score)
    except (TypeError, ValueError):
        return {"tier": "unknown", "label": "no score",
                "color": "ghost", "css_class": "badge-ghost"}
    if s >= 0.90:
        return {"tier": "auto", "label": f"Auto-approved ({s:.2f})",
                "color": "success", "css_class": "badge-success"}
    if s >= 0.60:
        return {"tier": "review", "label": f"Needs review ({s:.2f})",
                "color": "warning", "css_class": "badge-warning"}
    return {"tier": "attention", "label": f"Needs attention ({s:.2f})",
            "color": "error", "css_class": "badge-error"}


@app.route("/logistics/review")
def ui_logistics_review_queue():
    """Show all logistics documents with confidence < 0.90 that aren't reviewed yet."""
    if not os.path.exists(LOGISTICS_DB):
        return render_template("logistics/review_queue.html", mode="logistics",
                               docs=[], db_present=False)
    show_all = request.args.get("all") == "1"
    conn = get_connection(LOGISTICS_DB)
    try:
        if show_all:
            rows = conn.execute(
                """SELECT id, type, source_file, confidence, created_at, reviewed_at,
                          extracted_json
                   FROM documents
                   WHERE module = 'logistics'
                   ORDER BY confidence ASC NULLS LAST, id DESC LIMIT 200"""
            ).fetchall()
        else:
            rows = conn.execute(
                """SELECT id, type, source_file, confidence, created_at, reviewed_at,
                          extracted_json
                   FROM documents
                   WHERE module = 'logistics'
                     AND (confidence IS NULL OR confidence < 0.90)
                     AND reviewed_at IS NULL
                   ORDER BY confidence ASC NULLS LAST, id DESC LIMIT 200"""
            ).fetchall()
    finally:
        conn.close()

    docs = []
    for r in rows:
        try:
            ed = json.loads(r["extracted_json"] or "{}")
        except Exception:
            ed = {}
        tier = confidence_tier(r["confidence"])
        docs.append({
            "id":         r["id"],
            "type":       r["type"],
            "tan":        ed.get("tan_number"),
            "carrier":    ed.get("carrier"),
            "vessel":     ed.get("vessel_name"),
            "confidence": r["confidence"],
            "tier":       tier,
            "reviewed_at": r["reviewed_at"],
            "basename":   os.path.basename(r["source_file"] or ""),
            "created_at": r["created_at"],
        })
    return render_template("logistics/review_queue.html", mode="logistics",
                           docs=docs, db_present=True, show_all=show_all)


@app.route("/logistics/review/<int:doc_id>/approve", methods=["POST"])
def ui_logistics_review_approve(doc_id: int):
    """Mark a document as reviewed (operator-approved). Boosts to auto tier."""
    if not os.path.exists(LOGISTICS_DB):
        abort(404)
    conn = get_connection(LOGISTICS_DB)
    try:
        conn.execute(
            """UPDATE documents
               SET reviewed_at = datetime('now'),
                   reviewed_by = 'operator'
               WHERE id = ? AND module = 'logistics'""",
            (doc_id,),
        )
        try:
            conn.execute(
                """INSERT INTO audit_log (action, actor, entity_type, entity_id, timestamp)
                   VALUES ('review_approve', 'operator', 'documents', ?, datetime('now'))""",
                (str(doc_id),),
            )
        except Exception:
            pass
        conn.commit()
        return render_template("logistics/_edit_result.html", ok=True,
                               msg="Approved.", container_id=doc_id)
    except Exception as e:
        return render_template("logistics/_edit_result.html", ok=False,
                               msg=f"Failed: {e}", container_id=doc_id)
    finally:
        conn.close()


# P1-C re-extract diff + P1-A bounding-box helpers come from
# core.business.reconcile (imported with the other business helpers above).


def _file_sha256(path: str) -> str:
    """Recompute the same sha256 the processor uses, so we can clear cache."""
    import hashlib
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while chunk := f.read(8192):
            h.update(chunk)
    return h.hexdigest()


@app.route("/logistics/documents/<int:doc_id>/reextract", methods=["POST"])
def ui_logistics_document_reextract(doc_id: int):
    """Run extraction again on the stored source file, return diff modal.
    Does NOT save — operator must explicitly Accept."""
    if not os.path.exists(LOGISTICS_DB):
        abort(404)
    conn = get_connection(LOGISTICS_DB)
    try:
        row = conn.execute(
            "SELECT source_file, extracted_json FROM documents WHERE id = ? AND module='logistics'",
            (doc_id,),
        ).fetchone()
    finally:
        conn.close()
    if not row:
        abort(404, f"Document {doc_id} not found")

    source = row["source_file"]
    if not source or not os.path.isfile(source):
        return render_template(
            "logistics/_reextract_diff.html",
            ok=False,
            error=f"Source file not found on disk: {os.path.basename(source or '')}",
            doc_id=doc_id,
        )

    # Clear cache so the LLM actually runs again (otherwise we'd just get the old result)
    try:
        file_hash = _file_sha256(source)
        c2 = get_connection(LOGISTICS_DB)
        c2.execute("DELETE FROM extraction_cache WHERE file_hash = ?", (file_hash,))
        c2.commit()
        c2.close()
    except Exception as e:
        return render_template("logistics/_reextract_diff.html",
                               ok=False, error=f"Cache clear failed: {e}", doc_id=doc_id)

    # Run the pipeline synchronously (LLM call takes 10-30s)
    try:
        from core.pipeline.router import route_file
        job = route_file(source, "logistics")
    except Exception as e:
        return render_template("logistics/_reextract_diff.html",
                               ok=False, error=f"Re-extract crashed: {type(e).__name__}: {e}",
                               doc_id=doc_id)

    if job.status != "COMPLETED":
        return render_template(
            "logistics/_reextract_diff.html",
            ok=False,
            error=f"Re-extract failed: {job.error_message or 'unknown'}",
            doc_id=doc_id,
        )

    new_data = (job.result_data or {}).get("extracted_data") or {}
    try:
        old_data = json.loads(row["extracted_json"] or "{}")
    except Exception:
        old_data = {}
    diff = _compute_diff(old_data, new_data)

    return render_template(
        "logistics/_reextract_diff.html",
        ok=True,
        doc_id=doc_id,
        diff=diff,
        new_json=json.dumps(new_data, ensure_ascii=False),
        new_json_pretty=json.dumps(new_data, indent=2, ensure_ascii=False, default=str),
    )


@app.route("/logistics/documents/<int:doc_id>/reextract/accept", methods=["POST"])
def ui_logistics_document_reextract_accept(doc_id: int):
    """Persist the re-extracted JSON. Logs old vs new to audit_log."""
    if not os.path.exists(LOGISTICS_DB):
        abort(404)
    new_json = request.form.get("new_json", "")
    try:
        new_data = json.loads(new_json) if new_json else {}
    except json.JSONDecodeError as e:
        return render_template("logistics/_edit_result.html", ok=False,
                               msg=f"Invalid JSON: {e}", container_id=doc_id)

    conn = get_connection(LOGISTICS_DB)
    try:
        row = conn.execute(
            "SELECT extracted_json FROM documents WHERE id=? AND module='logistics'",
            (doc_id,),
        ).fetchone()
        if not row:
            abort(404)
        before = row["extracted_json"]

        conn.execute(
            "UPDATE documents SET extracted_json = ? WHERE id = ?",
            (json.dumps(new_data, ensure_ascii=False), doc_id),
        )
        # Audit log entry
        try:
            conn.execute(
                """INSERT INTO audit_log (action, actor, entity_type, entity_id,
                                          before_json, after_json, timestamp)
                   VALUES ('reextract', 'operator', 'documents', ?, ?, ?, datetime('now'))""",
                (str(doc_id), before, json.dumps(new_data, ensure_ascii=False)),
            )
        except Exception:
            pass  # audit log is best-effort
        conn.commit()
        return render_template("logistics/_edit_result.html", ok=True,
                               msg="Re-extraction accepted. Reload to see updated fields.",
                               container_id=doc_id)
    except Exception as e:
        return render_template("logistics/_edit_result.html", ok=False,
                               msg=f"Save failed: {e}", container_id=doc_id)
    finally:
        conn.close()


@app.route("/logistics/documents/<int:doc_id>", methods=["POST"])
def ui_logistics_document_update(doc_id: int):
    """Save edits — supports two modes:
       1. Path-based: form fields named like `f.tan_number`, `f.containers[0].size` —
          parse + write back into nested JSON.
       2. Raw mode: a single `raw_json` field — replaces the entire extracted_json.
    """
    if not os.path.exists(LOGISTICS_DB):
        abort(404)
    conn = get_connection(LOGISTICS_DB)
    try:
        row = conn.execute(
            "SELECT extracted_json FROM documents WHERE id = ? AND module='logistics'",
            (doc_id,),
        ).fetchone()
        if not row:
            abort(404)
        try:
            data = json.loads(row["extracted_json"] or "{}")
        except Exception:
            data = {}

        if "raw_json" in request.form and (request.form.get("raw_json") or "").strip():
            try:
                data = json.loads(request.form["raw_json"])
            except json.JSONDecodeError as e:
                return render_template("logistics/_edit_result.html",
                                       ok=False,
                                       msg=f"Invalid JSON: {e}",
                                       container_id=doc_id)
            n_changed = "ALL"
        else:
            n_changed = 0
            for key, value in request.form.items():
                if not key.startswith("f."):
                    continue
                path = key[2:]
                v = value.strip()
                # Try to coerce to int/float if it looks numeric
                if v == "":
                    v = None
                elif v.lower() in ("true", "false"):
                    v = (v.lower() == "true")
                else:
                    try:
                        v = int(v) if v.isdigit() or (v.startswith("-") and v[1:].isdigit()) else float(v)
                    except ValueError:
                        pass  # keep as string
                _set_nested(data, path, v)
                n_changed += 1

        conn.execute(
            "UPDATE documents SET extracted_json = ? WHERE id = ?",
            (json.dumps(data, ensure_ascii=False), doc_id),
        )
        conn.commit()
        return render_template("logistics/_edit_result.html",
                               ok=True,
                               msg=f"Saved — {n_changed} field(s) updated",
                               container_id=doc_id)
    except Exception as e:
        return render_template("logistics/_edit_result.html",
                               ok=False, msg=f"DB error: {e}",
                               container_id=doc_id)
    finally:
        conn.close()


# ─── File serve for shipments (for split-view PDF) ────────────────────────────
@app.route("/files/logistics/shipment/<int:shipment_id>")
def ui_serve_logistics_shipment_file(shipment_id: int):
    """Serve the source PDF for a shipment (looked up from shipments.source_file)."""
    if not os.path.exists(LOGISTICS_DB):
        abort(404)
    conn = get_connection(LOGISTICS_DB)
    try:
        row = conn.execute(
            "SELECT source_file FROM shipments WHERE id = ?", (shipment_id,)
        ).fetchone()
    finally:
        conn.close()
    if not row or not row["source_file"]:
        abort(404, "No source file recorded for this shipment")
    path = row["source_file"]
    if not os.path.isfile(path):
        # Try looking in trial_files and data/input as fallback
        base = os.path.basename(path)
        for search_dir in ("data/trial_files", "data/input", "fresh documents/logistics documents"):
            candidate = os.path.join(ROOT_DIR, search_dir, base)
            if os.path.isfile(candidate):
                path = candidate
                break
        else:
            abort(404, f"Source file not found: {base}")
    mime, _ = mimetypes.guess_type(path)
    return send_file(path, mimetype=mime or "application/pdf")


@app.route("/logistics/containers/<int:container_id>", methods=["GET"])
def ui_logistics_container_detail(container_id: int):
    """Split-view: original PDF on the left, editable form + Copy JSON on the right."""
    row = _fetch_container_row(container_id)
    if not row:
        abort(404, f"Container {container_id} not found")
    # Build the extracted JSON for display + copy button
    extracted_json = json.dumps(
        {k: v for k, v in row.items() if v is not None},
        indent=2,
        default=str,
        ensure_ascii=False,
    )
    # Find shipment_id for the file-serve route
    conn = get_connection(LOGISTICS_DB)
    try:
        ship = conn.execute(
            "SELECT id, source_file FROM shipments WHERE id = "
            "(SELECT shipment_id FROM containers WHERE id = ?)",
            (container_id,),
        ).fetchone()
        shipment_id = ship["id"] if ship else None
        source_file = ship["source_file"] if ship else None
        source_basename = os.path.basename(source_file) if source_file else None
        ext = os.path.splitext(source_file or "")[1].lower()
        # P2-C: fetch full shipment row for demurrage calc
        shipment_full = conn.execute(
            "SELECT * FROM shipments WHERE id = ?",
            (shipment_id,)
        ).fetchone() if shipment_id else None
    finally:
        conn.close()
    has_file = bool(source_file)

    # P2-C: Compute demurrage risk. N5: enrich shipment dict with free_days
    # from the LLM-extracted demurrage clause if available — overrides the
    # carrier default for this specific shipment's actual contract.
    shipment_dict = dict(shipment_full) if shipment_full else {}
    if shipment_dict.get("tan"):
        fd = _free_days_from_documents(LOGISTICS_DB, shipment_dict["tan"])
        if fd is not None:
            shipment_dict["free_days"] = fd
    dd_info = _demurrage_info(dict(row), shipment_dict)

    return render_template(
        "logistics/container_detail.html",
        mode="logistics",
        c=row,
        fields=EDITABLE_FIELDS,
        extracted_json=extracted_json,
        shipment_id=shipment_id,
        source_basename=source_basename,
        has_file=has_file,
        ext=ext,
        dd=dd_info,
    )


@app.route("/logistics/containers/<int:container_id>", methods=["POST"])
def ui_logistics_container_update(container_id: int):
    """HTMX inline-edit for a container row."""
    if not os.path.exists(LOGISTICS_DB):
        abort(404)
    updates: dict = {}
    for name, kind in EDITABLE_FIELDS.items():
        if name not in request.form:
            continue
        raw = request.form.get(name, "").strip()
        if raw == "":
            updates[name] = None
        elif kind == "number":
            try:
                updates[name] = float(raw)
            except ValueError:
                updates[name] = None
        else:
            updates[name] = raw
    if not updates:
        return render_template("logistics/_edit_result.html",
                               ok=False, msg="No fields to update",
                               container_id=container_id)
    conn = get_connection(LOGISTICS_DB)
    try:
        cols = ", ".join(f"{k} = ?" for k in updates)
        params = list(updates.values()) + [container_id]
        conn.execute(
            f"UPDATE containers SET {cols}, modified_at = datetime('now') WHERE id = ?",
            params,
        )
        conn.commit()
        return render_template("logistics/_edit_result.html",
                               ok=True, msg=f"Saved — {len(updates)} field(s) updated.",
                               container_id=container_id)
    except Exception as e:
        return render_template("logistics/_edit_result.html",
                               ok=False, msg=f"DB error: {e}",
                               container_id=container_id)
    finally:
        conn.close()


@app.route("/logistics/edit/<int:container_id>", methods=["GET"])
def ui_logistics_edit(container_id: int):
    row = _fetch_container_row(container_id)
    if not row:
        abort(404, f"Container {container_id} not found")
    return render_template(
        "logistics/edit.html",
        mode="logistics",
        c=row,
        fields=EDITABLE_FIELDS,
    )


# ─── Export (P3.3) ────────────────────────────────────────────────────────────
import csv
import io
from flask import Response

# Export column registry + query builder live in core.business.exports.
from core.business.exports import EXPORT_COLUMNS, run_query as _run_export_query


def _export_query(picked_aliases: set[str], filters: dict):
    """Thin wrapper that injects LOGISTICS_DB."""
    return _run_export_query(LOGISTICS_DB, picked_aliases, filters)


@app.route("/logistics/swimlane")
def ui_logistics_swimlane():
    """P2-E: Container status swimlane — visual pipeline view."""
    STAGES = ["BOOKED", "IN_TRANSIT", "ARRIVED", "DELIVERED", "RESTITUTED", "CLOSED"]

    if not os.path.exists(LOGISTICS_DB):
        return render_template("logistics/swimlane.html", mode="logistics",
                               stages=STAGES, columns={}, db_present=False)

    conn = get_connection(LOGISTICS_DB)
    try:
        rows = conn.execute(
            """SELECT c.id, c.container_number, c.size, c.statut_container,
                      s.tan, s.compagnie_maritime, s.port, s.eta, s.etd,
                      c.date_livraison, c.date_depotement
               FROM containers c
               JOIN shipments s ON s.id = c.shipment_id
               ORDER BY c.id DESC"""
        ).fetchall()
    finally:
        conn.close()

    # Bin containers by status
    columns = {stage: [] for stage in STAGES}
    other = []
    for r in rows:
        status = (r["statut_container"] or "").strip().upper()
        # Normalise
        if status in ("EN_ROUTE", "EN TRANSIT"):
            status = "IN_TRANSIT"
        if status in columns:
            columns[status].append(dict(r))
        else:
            other.append(dict(r))

    if other:
        columns.setdefault("OTHER", []).extend(other)

    # N5: build a TAN → free_days cache once so we don't open a connection
    # per card. Empty cache when no TAN has document-extracted free_days.
    distinct_tans = {c.get("tan") for stage in columns.values() for c in stage if c.get("tan")}
    fd_cache: dict[str, int] = {}
    for tan in distinct_tans:
        fd = _free_days_from_documents(LOGISTICS_DB, tan)
        if fd is not None:
            fd_cache[tan] = fd

    # Compute D&D risk per card, using extracted free_days when available.
    for stage in list(columns.keys()):
        for card in columns[stage]:
            shipment = {
                "eta": card.get("eta"),
                "compagnie_maritime": card.get("compagnie_maritime"),
            }
            tan = card.get("tan")
            if tan and tan in fd_cache:
                shipment["free_days"] = fd_cache[tan]
            card["dd"] = _demurrage_info(card, shipment)

    return render_template("logistics/swimlane.html", mode="logistics",
                           stages=STAGES, columns=columns, db_present=True)


# ─── P4-A: Global search (hybrid: SQL LIKE + ChromaDB semantic) ─────────────

@app.route("/search")
def ui_global_search():
    q = (request.args.get("q") or "").strip()
    mode_arg = (request.args.get("mode") or "").strip()
    if not q:
        return render_template("global_search.html", q="", results=[], mode=mode_arg or None)

    results = []
    # Logistics keyword + semantic
    if os.path.exists(LOGISTICS_DB) and (mode_arg in ("", "logistics")):
        conn = get_connection(LOGISTICS_DB)
        try:
            rows = conn.execute(
                """SELECT id, type, source_file, extracted_json
                   FROM documents
                   WHERE module = 'logistics' AND (
                         extracted_json LIKE ? OR source_file LIKE ?)
                   ORDER BY id DESC LIMIT 20""",
                (f"%{q}%", f"%{q}%"),
            ).fetchall()
        finally:
            conn.close()
        seen = set()
        for r in rows:
            try:
                ed = json.loads(r["extracted_json"] or "{}")
            except Exception:
                ed = {}
            seen.add(r["id"])
            results.append({
                "module": "logistics",
                "url": f"/logistics/documents/{r['id']}",
                "title": ed.get("tan_number") or os.path.basename(r["source_file"] or "") or f"doc {r['id']}",
                "subtitle": (r["type"] or "") + " · " + (ed.get("vessel_name") or ed.get("carrier") or ""),
                "match": "keyword",
            })
        # Add top semantic hits not already in keyword results
        for did in _semantic_search_ids(q, "logistics", n=8):
            if did in seen:
                continue
            results.append({"module": "logistics",
                            "url": f"/logistics/documents/{did}",
                            "title": f"doc {did}",
                            "subtitle": "matched by meaning",
                            "match": "semantic"})

    # Travel keyword
    if os.path.exists(TRAVEL_DB) and (mode_arg in ("", "travel")):
        conn = get_connection(TRAVEL_DB)
        try:
            persons = conn.execute(
                """SELECT id, full_name, nationality FROM persons
                   WHERE full_name LIKE ? OR nationality LIKE ?
                   ORDER BY id DESC LIMIT 10""",
                (f"%{q}%", f"%{q}%"),
            ).fetchall()
            fams = conn.execute(
                """SELECT id, family_name, case_reference FROM families
                   WHERE family_name LIKE ? OR case_reference LIKE ?
                   ORDER BY id DESC LIMIT 10""",
                (f"%{q}%", f"%{q}%"),
            ).fetchall()
            tdocs = conn.execute(
                """SELECT id, doc_type, doc_number, person_id FROM documents_travel
                   WHERE doc_number LIKE ? ORDER BY id DESC LIMIT 10""",
                (f"%{q}%",),
            ).fetchall()
        finally:
            conn.close()
        for p in persons:
            results.append({"module": "travel",
                            "url": f"/travel/persons/{p['id']}",
                            "title": p["full_name"] or "Person #" + str(p["id"]),
                            "subtitle": p["nationality"] or "person",
                            "match": "keyword"})
        for f in fams:
            results.append({"module": "travel",
                            "url": f"/travel/families/{f['id']}",
                            "title": f["family_name"] or "Family #" + str(f["id"]),
                            "subtitle": f["case_reference"] or "family",
                            "match": "keyword"})
        for d in tdocs:
            results.append({"module": "travel",
                            "url": f"/travel/documents/{d['id']}",
                            "title": d["doc_number"] or "—",
                            "subtitle": d["doc_type"] or "doc",
                            "match": "keyword"})
    return render_template("global_search.html", q=q, results=results, mode=mode_arg or None)


# ─── P5-A: Logistics analytics ──────────────────────────────────────────────

# Chart data extracted to core.business.charts (TD8 phase 1).
from core.business.charts import (
    logistics_chart_data as _logistics_chart_data_impl,
    travel_chart_data as _travel_chart_data_impl,
)


@app.route("/logistics/analytics")
def ui_logistics_analytics():
    return render_template(
        "logistics/analytics.html",
        mode="logistics",
        charts=_logistics_chart_data_impl(LOGISTICS_DB),
        stats=_logistics_stats(),
        db_present=os.path.exists(LOGISTICS_DB),
    )


# ─── P5-B: Travel analytics ─────────────────────────────────────────────────

@app.route("/travel/analytics")
def ui_travel_analytics():
    return render_template(
        "travel/analytics.html",
        mode="travel",
        charts=_travel_chart_data_impl(TRAVEL_DB),
        db_present=os.path.exists(TRAVEL_DB),
    )


@app.route("/logistics/export", methods=["GET"])
def ui_logistics_export():
    return render_template(
        "logistics/export.html",
        mode="logistics",
        columns=EXPORT_COLUMNS,
    )


@app.route("/logistics/sheet", methods=["GET"])
def ui_logistics_sheet():
    """Interactive sheet viewer — sort + filter + customize columns + export."""
    if not os.path.exists(LOGISTICS_DB):
        return render_template(
            "logistics/sheet.html",
            mode="logistics",
            columns=EXPORT_COLUMNS,
            rows=[],
            labels=[],
            filters={"status": None, "carrier": None, "tan": None},
            db_present=False,
        )
    filters = {k: (request.args.get(k) or "").strip() or None
               for k in ("status", "carrier", "tan")}
    # Always pull ALL columns at the API level — the customizer hides/orders
    # client-side so the user can toggle without server round-trips.
    all_aliases = {a for a, _ in EXPORT_COLUMNS}
    labels, rows = _export_query(all_aliases, filters)
    # Build (alias, label) list in the EXPORT_COLUMNS order so the JS knows
    # the canonical mapping
    return render_template(
        "logistics/sheet.html",
        mode="logistics",
        columns=EXPORT_COLUMNS,
        labels=labels,
        rows=[list(r) for r in rows],
        filters=filters,
        db_present=True,
    )


@app.route("/logistics/export.csv")
def ui_logistics_export_csv():
    if not os.path.exists(LOGISTICS_DB):
        abort(404, "Logistics DB not found")
    picked = {c.strip() for c in request.args.get("cols", "").split(",") if c.strip()}
    filters = {k: request.args.get(k, "").strip() or None for k in ("status","carrier","tan")}
    labels, rows = _export_query(picked, filters)

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(labels)
    for r in rows:
        w.writerow(["" if v is None else v for v in r])
    return Response(
        "﻿" + buf.getvalue(),  # BOM so Excel opens UTF-8 cleanly
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="logistics_export.csv"'},
    )


@app.route("/logistics/export.xlsx")
def ui_logistics_export_xlsx():
    if not os.path.exists(LOGISTICS_DB):
        abort(404, "Logistics DB not found")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        abort(500, "openpyxl not installed")

    picked = {c.strip() for c in request.args.get("cols", "").split(",") if c.strip()}
    filters = {k: request.args.get(k, "").strip() or None for k in ("status","carrier","tan")}
    labels, rows = _export_query(picked, filters)

    wb = Workbook()
    ws = wb.active
    ws.title = "Containers"
    ws.append(labels)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6366F1")
    for r in rows:
        ws.append(["" if v is None else v for v in r])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, label in enumerate(labels, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(12, min(40, len(label) + 4))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="logistics_export.xlsx"'},
    )


@app.route("/logistics/edit/<int:container_id>", methods=["POST"])
def ui_logistics_edit_post(container_id: int):
    if not os.path.exists(LOGISTICS_DB):
        abort(404)

    # Coerce + collect fields that came back in the form
    updates: dict = {}
    for name, kind in EDITABLE_FIELDS.items():
        if name not in request.form:
            continue
        raw = request.form.get(name, "").strip()
        if raw == "":
            updates[name] = None
        elif kind == "number":
            try:
                updates[name] = float(raw)
            except ValueError:
                updates[name] = None
        else:
            updates[name] = raw

    if not updates:
        return render_template(
            "logistics/_edit_result.html",
            ok=False, msg="No fields to update", container_id=container_id,
        )

    conn = get_connection(LOGISTICS_DB)
    try:
        cols = ", ".join(f"{k} = ?" for k in updates.keys())
        params = list(updates.values()) + [container_id]
        conn.execute(
            f"UPDATE containers SET {cols}, modified_at = datetime('now') WHERE id = ?",
            params,
        )
        conn.commit()
        return render_template(
            "logistics/_edit_result.html",
            ok=True,
            msg=f"Updated {len(updates)} field(s)",
            container_id=container_id,
        )
    except Exception as e:
        return render_template(
            "logistics/_edit_result.html",
            ok=False, msg=f"DB error: {e}", container_id=container_id,
        )
    finally:
        conn.close()


# ─── Entry Point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("BRUNS_API_PORT", 7845))

    # Run pending migrations on every start. Safe + idempotent.
    try:
        from core.storage.migrations import run_migrations
        for db in (LOGISTICS_DB, TRAVEL_DB):
            if os.path.exists(db):
                applied = run_migrations(db)
                if applied:
                    log.info("migrations applied to %s: %s", os.path.basename(db), applied)
    except Exception as _mig_err:
        log.warning("migration runner failed: %s", _mig_err)

    print(f"\n{'='*60}")
    print(f"  BRUNs Document Intelligence")
    print(f"  Running at: http://localhost:{port}")
    print(f"  UI:         http://localhost:{port}/")
    print(f"  Power BI:   http://localhost:{port}/api/logistics/shipments_full")
    print(f"{'='*60}\n")
    app.run(host="0.0.0.0", port=port, debug=False)
