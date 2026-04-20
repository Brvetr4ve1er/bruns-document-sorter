"""
BRUNs Logistics Dashboard — Complete Unified App
All features from both app.py and app_new.py merged into the glass bottom-nav design.
Pages: Dashboard · Containers · Edit · Processing · Export · Settings/Logs
"""
import os, sys, json, sqlite3, shutil, tempfile
from datetime import datetime, date

import streamlit as st
import pandas as pd

sys.path.insert(0, os.path.dirname(__file__))

import settings_store
from db.database import init_db, export_to_csv, update_container
from ui.styles import inject_global_css, COLORS, STATUS_STYLE
from ui.layout import inject_layout_css, initialize_layout_state, render_page_header
from ui.components import (
    render_filter_bar, apply_filters, render_filter_summary,
    render_document_grid, render_document_viewer, render_bulk_actions_bar,
)
from ui.pages import render_processing_page, render_export_page, render_settings_page

# ── Page config — MUST be first ───────────────────────────────────────────────
st.set_page_config(
    page_title="BRUNs Logistics",
    page_icon="🚢",
    layout="wide",
    initial_sidebar_state="collapsed",
)

inject_global_css()
inject_layout_css()

if "settings" not in st.session_state:
    st.session_state.settings = settings_store.load()
settings_store.apply_to_config(st.session_state.settings)
initialize_layout_state()
init_db()

# Apply saved UI settings (Ollama overrides, etc.)
try:
    _p = os.path.join(os.path.dirname(__file__), "data", "ui_settings.json")
    if os.path.exists(_p):
        with open(_p) as f:
            _ui = json.load(f)
        import config as _cfg
        if _ui.get("ollama_url"):   _cfg.OLLAMA_URL   = _ui["ollama_url"].rstrip("/") + "/api/generate"
        if _ui.get("ollama_model"): _cfg.OLLAMA_MODEL = _ui["ollama_model"]
        if _ui.get("ollama_timeout"): _cfg.OLLAMA_TIMEOUT = _ui["ollama_timeout"]
except Exception:
    pass


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _conn():
    import config
    c = sqlite3.connect(config.DB_PATH)
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA foreign_keys = ON")
    return c

def fetch_containers_view() -> pd.DataFrame:
    conn = _conn()
    df = pd.read_sql_query("""
        SELECT c.id AS container_id, s.id AS shipment_id,
            c.container_number   AS "N° Container",
            s.tan                AS "N° TAN",
            s.item_description   AS "Item",
            s.compagnie_maritime AS "Compagnie maritime",
            s.port               AS "Port",
            s.transitaire        AS "Transitaire",
            s.etd                AS "Date shipment",
            s.eta                AS "Date accostage",
            c.statut_container   AS "Statut Container",
            c.size               AS "Container size",
            c.seal_number        AS "N° Seal",
            c.date_livraison     AS "Date livraison",
            c.site_livraison     AS "Site livraison",
            c.date_depotement    AS "Date dépotement",
            c.date_restitution   AS "Date restitution",
            c.restitue_chauffeur AS "Chauffeur restitution",
            c.livre_chauffeur    AS "Chauffeur livraison",
            c.commentaire        AS "Commentaire",
            s.source_file        AS "Source",
            c.created_at         AS "Créé le"
        FROM containers c JOIN shipments s ON s.id = c.shipment_id
        ORDER BY c.id DESC
    """, conn)
    conn.close()
    return df

def fetch_stats() -> dict:
    conn = _conn()
    s  = conn.execute("SELECT COUNT(*) FROM shipments").fetchone()[0]
    c  = conn.execute("SELECT COUNT(*) FROM containers").fetchone()[0]
    bk = conn.execute("SELECT COUNT(*) FROM shipments WHERE status='BOOKED'").fetchone()[0]
    tr = conn.execute("SELECT COUNT(*) FROM shipments WHERE status='IN_TRANSIT'").fetchone()[0]
    bc = conn.execute("SELECT compagnie_maritime, COUNT(*) c FROM shipments WHERE compagnie_maritime IS NOT NULL GROUP BY compagnie_maritime ORDER BY c DESC").fetchall()
    conn.close()
    return {"shipments": s, "containers": c, "booked": bk, "in_transit": tr,
            "by_carrier": [(r[0], r[1]) for r in bc]}

def process_pdf_file(pdf_path: str, original_name: str = None) -> dict:
    from parsers.pdf_extractor import extract_text
    from agents.parser_agent import parse_document
    from utils.validator import parse_and_validate
    from utils.logger import log_result
    filename = original_name or os.path.basename(pdf_path)
    try:
        text = extract_text(pdf_path)
    except Exception as e:
        return {"file": filename, "status": "EXTRACTION_FAILED", "error": str(e)}
    if not text.strip():
        return {"file": filename, "status": "EMPTY_PDF", "error": "No text extracted"}
    raw, llm_err = parse_document(text)
    if llm_err:
        return {"file": filename, "status": "LLM_FAILED", "error": llm_err}
    model, raw_dict, val_err = parse_and_validate(raw)
    if val_err or model is None:
        return {"file": filename, "status": "VALIDATION_FAILED", "error": val_err, "raw": raw_dict}
    try:
        from db.database import upsert_shipment
        action, sid = upsert_shipment(model, source_file=filename)
    except Exception as e:
        return {"file": filename, "status": "DB_ERROR", "error": str(e)}
    log_result(filename, raw_dict, True, action)
    return {"file": filename, "status": "OK", "db_action": action,
            "shipment_id": sid, "data": raw_dict,
            "container_count": len(model.containers)}

def _parse_date(s):
    if not s: return None
    if isinstance(s, date): return s
    try: return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()
    except Exception: return None

def _date_str(d):
    if d is None: return None
    if isinstance(d, (datetime, date)): return d.strftime("%Y-%m-%d")
    return str(d)

def _get_log_files() -> list:
    logs_dir = st.session_state.settings.get("logs_dir", "")
    if not os.path.isdir(logs_dir): return []
    return sorted([f for f in os.listdir(logs_dir) if f.endswith(".json")], reverse=True)

def _grid_cols():
    try:
        p = os.path.join(os.path.dirname(__file__), "data", "ui_settings.json")
        if os.path.exists(p):
            with open(p) as f: return json.load(f).get("grid_columns", 3)
    except Exception: pass
    return 3

def _bulk_csv(df): return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")

def _bulk_xlsx(df):
    import io
    from openpyxl.styles import PatternFill, Font, Alignment
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Logistics")
        ws = w.sheets["Logistics"]
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max((len(str(c.value or "")) for c in col), default=8) + 4, 40)
        fill = PatternFill(start_color="1B2341", end_color="1B2341", fill_type="solid")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(color="E6E9F2", bold=True)
            cell.alignment = Alignment(horizontal="center")
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# BOTTOM NAVBAR — 6 pages
# ═══════════════════════════════════════════════════════════════════════════════

def render_navbar():
    page = st.session_state.get("current_page", "dashboard")

    NAV = [
        ("dashboard",   "📊 Dashboard"),
        ("containers",  "📋 Containers"),
        ("edit",        "✏️ Edit"),
        ("processing",  "⚙️ Processing"),
        ("export",      "📤 Export"),
        ("settings",    "🛠️ Settings"),
    ]

    cols = st.columns(len(NAV))
    for col, (key, label) in zip(cols, NAV):
        with col:
            is_active = page == key
            # Inline style to highlight active button
            if is_active:
                st.markdown(f"""
                <style>
                div[data-testid="column"]:nth-child({list(dict(NAV).keys()).index(key)+1}) button {{
                    background: linear-gradient(135deg, #6366F1, #8B5CF6) !important;
                    color: white !important;
                    border-color: #6366F1 !important;
                }}
                </style>
                """, unsafe_allow_html=True)
            if st.button(label, key=f"nav_{key}", use_container_width=True):
                st.session_state.current_page = key
                st.session_state.show_viewer = False
                st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE ROUTER
# ═══════════════════════════════════════════════════════════════════════════════

page = st.session_state.get("current_page", "dashboard")


# ── DASHBOARD ─────────────────────────────────────────────────────────────────
if page == "dashboard":
    render_page_header("Dashboard", "Filter and browse all processed containers", "📊")

    try:
        stats = fetch_stats()
        st.markdown('<div data-aos="fade-up" data-aos-delay="80">', unsafe_allow_html=True)
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("📦 Shipments",  f"{stats['shipments']:,}")
        c2.metric("🗃️ Containers", f"{stats['containers']:,}")
        c3.metric("🟢 Booked",     f"{stats['booked']:,}")
        c4.metric("🚢 In Transit", f"{stats['in_transit']:,}")
        st.markdown('</div>', unsafe_allow_html=True)
    except Exception as e:
        st.warning(f"Stats unavailable: {e}")

    st.divider()

    try:
        df = fetch_containers_view()
    except Exception as e:
        st.error(f"❌ {e}")
        df = pd.DataFrame()

    if df.empty:
        st.info("📭 No containers yet. Go to **⚙️ Processing** to import PDF files.")
    else:
        filters = render_filter_bar(df)
        filtered_df = apply_filters(df, filters)
        render_filter_summary(filters, len(df), len(filtered_df))
        st.divider()
        render_document_grid(filtered_df, columns=_grid_cols())
        render_document_viewer()

        # Bulk export wiring
        trigger = st.session_state.pop("bulk_export_trigger", None)
        if trigger:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            export_cols = [c for c in [
                "N° Container","N° TAN","Item","Compagnie maritime","Port","Transitaire",
                "Date shipment","Date accostage","Statut Container","Container size",
                "N° Seal","Date livraison","Site livraison","Date dépotement","Date restitution",
                "Commentaire","Source"
            ] if c in filtered_df.columns]
            edf = filtered_df[export_cols]
            if trigger == "csv":
                st.download_button("⬇️ Download CSV", _bulk_csv(edf),
                    f"bruns_filtered_{ts}.csv", "text/csv")
            elif trigger == "xlsx":
                st.download_button("⬇️ Download Excel", _bulk_xlsx(edf),
                    f"bruns_filtered_{ts}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.divider()
        render_bulk_actions_bar(len(filtered_df))


# ── CONTAINERS TABLE ──────────────────────────────────────────────────────────
elif page == "containers":
    render_page_header("Containers", "Full table view — sort, filter, inspect every field", "📋")

    df = fetch_containers_view()
    if df.empty:
        st.info("📭 No containers yet. Go to **⚙️ Processing** to import PDF files.")
        st.stop()

    # Quick stats
    stats = fetch_stats()
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📦 Shipments",  stats["shipments"])
    c2.metric("🗃️ Containers", stats["containers"])
    c3.metric("🟢 Booked",     stats["booked"])
    c4.metric("🚢 In Transit", stats["in_transit"])
    st.divider()

    # Carrier chart
    if stats["by_carrier"]:
        with st.expander("📊 By Carrier", expanded=False):
            bc_df = pd.DataFrame(stats["by_carrier"], columns=["Carrier", "Shipments"])
            st.bar_chart(bc_df.set_index("Carrier"))

    # Filters
    with st.expander("🔍 Filters", expanded=True):
        fc1, fc2, fc3, fc4 = st.columns(4)
        carriers = sorted([x for x in df["Compagnie maritime"].dropna().unique()])
        f_car  = fc1.multiselect("Carrier",  carriers, default=carriers, key="ct_car")
        sizes  = sorted([x for x in df["Container size"].dropna().unique()])
        f_size = fc2.multiselect("Size",     sizes,    default=sizes,    key="ct_size")
        stats_ = sorted([x for x in df["Statut Container"].dropna().unique()])
        f_stat = fc3.multiselect("Statut",   stats_,   default=stats_,   key="ct_stat")
        f_txt  = fc4.text_input("🔎 Search", placeholder="TAN, container, item…", key="ct_txt")

    filt = df.copy()
    if f_car:  filt = filt[filt["Compagnie maritime"].isin(f_car)]
    if f_size: filt = filt[filt["Container size"].isin(f_size)]
    if f_stat: filt = filt[filt["Statut Container"].isin(f_stat)]
    if f_txt:
        p = f_txt.lower()
        filt = filt[
            filt["N° Container"].fillna("").str.lower().str.contains(p) |
            filt["N° TAN"].fillna("").str.lower().str.contains(p) |
            filt["Item"].fillna("").str.lower().str.contains(p)
        ]

    st.caption(f"Showing **{len(filt):,}** of **{len(df):,}** containers")

    display_cols = [c for c in filt.columns if c not in ("container_id", "shipment_id")]
    st.dataframe(
        filt[display_cols],
        use_container_width=True, hide_index=True, height=600,
        column_config={
            "Date shipment":    st.column_config.DateColumn("ETD"),
            "Date accostage":   st.column_config.DateColumn("ETA"),
            "Date livraison":   st.column_config.DateColumn(),
            "Date dépotement":  st.column_config.DateColumn(),
            "Date restitution": st.column_config.DateColumn(),
            "Créé le":          st.column_config.DatetimeColumn(),
        },
    )

    # Shipments sub-tab
    st.divider()
    with st.expander("🚚 Shipments view"):
        conn = _conn()
        sdf = pd.read_sql_query("""
            SELECT s.id, s.tan, s.vessel, s.compagnie_maritime AS carrier,
                   s.transitaire, s.item_description AS item,
                   s.etd, s.eta, s.status, s.document_type, s.source_file,
                   (SELECT COUNT(*) FROM containers c WHERE c.shipment_id=s.id) AS containers
            FROM shipments s ORDER BY s.id DESC
        """, conn)
        conn.close()
        st.dataframe(sdf, use_container_width=True, hide_index=True, height=400)

        st.markdown("**Delete a shipment** (cascades to containers):")
        if not sdf.empty:
            dc1, dc2 = st.columns([1, 3])
            del_id = dc1.selectbox("Shipment ID", sdf["id"].tolist(), key="del_sid")
            if dc2.button("⚠️ Delete", key="del_ship_btn"):
                conn2 = _conn()
                with conn2: conn2.execute("DELETE FROM shipments WHERE id=?", (int(del_id),))
                conn2.close()
                st.toast(f"Shipment #{del_id} deleted", icon="🗑")
                st.rerun()


# ── EDIT CONTAINER ────────────────────────────────────────────────────────────
elif page == "edit":
    render_page_header("Edit Container", "Update delivery, demurrage, customs and driver info", "✏️")

    df = fetch_containers_view()
    if df.empty:
        st.info("📭 No containers yet. Go to **⚙️ Processing** to import PDF files.")
        st.stop()

    labels = [
        f"#{r.container_id}  ·  {r['N° Container']}  ·  {r['N° TAN'] or '—'}  ·  {r['Statut Container'] or '—'}"
        for r in df.itertuples()
    ]
    idx = st.selectbox("Select container", range(len(labels)), format_func=lambda i: labels[i])
    row = df.iloc[idx]
    cid = int(row["container_id"])

    conn = _conn()
    c = dict(conn.execute("SELECT * FROM containers WHERE id=?", (cid,)).fetchone())
    conn.close()

    # Header card
    status = c.get("statut_container") or "En transit"
    si = STATUS_STYLE.get(status, ("rgba(99,102,241,0.15)", "#6366F1", "📦"))
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,rgba(20,26,46,0.7),rgba(27,35,65,0.5));
                backdrop-filter:blur(12px); border:1px solid rgba(99,102,241,.25);
                border-radius:16px; padding:20px 24px; margin-bottom:24px;
                display:flex; gap:20px; align-items:center;">
        <div style="font-size:2.2rem;">📦</div>
        <div>
            <div style="font-size:1.25rem; font-weight:800; color:#E6E9F2;">{c['container_number']}</div>
            <div style="color:#8C95B3; font-size:0.85rem; margin-top:4px;">
                Shipment #{c['shipment_id']} · {row['N° TAN'] or 'no TAN'} · {row['Compagnie maritime'] or '—'}
            </div>
        </div>
        <div style="margin-left:auto;">
            <span style="background:{si[0]};color:{si[1]};padding:6px 14px;border-radius:20px;
                         font-size:0.8rem;font-weight:700;border:1px solid {si[1]}44;">
                {si[2]} {status}
            </span>
        </div>
    </div>
    """, unsafe_allow_html=True)

    with st.form("edit_container_form"):
        tab_a, tab_b, tab_c, tab_d = st.tabs(
            ["📦 Status & Delivery", "🚛 Transport", "💶 Demurrage & Billing", "🛃 Douane"]
        )

        with tab_a:
            ca, cb, cc = st.columns(3)
            STATUTS = ["Réservé","En transit","Arrivé","Livré","Dépoté","Restitué"]
            statut = ca.selectbox("Statut Container", STATUTS,
                index=STATUTS.index(status) if status in STATUTS else 1)
            SIZES = ["40 feet","20 feet","40 feet refrigerated","20 feet refrigerated"]
            cur_size = c.get("size") or "40 feet"
            size = cb.selectbox("Container size", SIZES,
                index=SIZES.index(cur_size) if cur_size in SIZES else 0)
            SITES = ["","Rouiba","Boudouaou","Other"]
            cur_site = c.get("site_livraison") or ""
            site = cc.selectbox("Site livraison", SITES,
                index=SITES.index(cur_site) if cur_site in SITES else 0)

            cd, ce, cf = st.columns(3)
            date_liv  = cd.date_input("Date livraison",  value=_parse_date(c.get("date_livraison")))
            date_dep  = ce.date_input("Date dépotement", value=_parse_date(c.get("date_depotement")))
            date_rest = cf.date_input("Date restitution", value=_parse_date(c.get("date_restitution")))

        with tab_b:
            cg, ch = st.columns(2)
            livre_cam = cg.text_input("🚛 Livré par (Camion)",    value=c.get("livre_camion") or "")
            livre_ch  = cg.text_input("👤 Livré par (Chauffeur)", value=c.get("livre_chauffeur") or "")
            rest_cam  = ch.text_input("🚛 Restitué par (Camion)",    value=c.get("restitue_camion") or "")
            rest_ch   = ch.text_input("👤 Restitué par (Chauffeur)", value=c.get("restitue_chauffeur") or "")
            centre    = st.text_input("🏢 Centre de restitution", value=c.get("centre_restitution") or "")

        with tab_c:
            ci, cj, ck = st.columns(3)
            date_debut_sur  = ci.date_input("Date début Surestarie",       value=_parse_date(c.get("date_debut_surestarie")))
            date_rest_est   = cj.date_input("Date restitution estimative",  value=_parse_date(c.get("date_restitution_estimative")))
            taux = ck.number_input("Taux de change", value=float(c.get("taux_de_change") or 0), step=0.01)

            cl, cm, cn = st.columns(3)
            jours_est     = cl.number_input("Jours surestarie estimés",  value=int(c.get("nbr_jours_surestarie_estimes") or 0), step=1)
            jours_douane  = cm.number_input("Jours perdu en douane",     value=int(c.get("nbr_jours_perdu_douane") or 0), step=1)
            jours_facture = cn.number_input("Jours surestarie facturés", value=int(c.get("nbr_jour_surestarie_facture") or 0), step=1)

            co_, cp_, cq_ = st.columns(3)
            facture_check = co_.selectbox("Montant facturé (check)", ["No","Yes"],
                index=0 if (c.get("montant_facture_check") or "No") == "No" else 1)
            montant   = cp_.number_input("Montant facturé (DA)", value=float(c.get("montant_facture_da") or 0), step=1.0)
            n_facture = cq_.text_input("N° Facture compagnie",   value=c.get("n_facture_cm") or "")

        with tab_d:
            cr_, cs_ = st.columns(2)
            date_decl = cr_.date_input("Date déclaration douane",  value=_parse_date(c.get("date_declaration_douane")))
            date_lib  = cs_.date_input("Date libération douane",   value=_parse_date(c.get("date_liberation_douane")))
            comment   = st.text_area("💬 Commentaire", value=c.get("commentaire") or "", height=100)

        submitted = st.form_submit_button("💾 Save changes", type="primary", use_container_width=True)

    if submitted:
        update_container(cid, {
            "statut_container":            statut,
            "size":                        size,
            "site_livraison":              site or None,
            "date_livraison":              _date_str(date_liv),
            "date_depotement":             _date_str(date_dep),
            "date_restitution":            _date_str(date_rest),
            "livre_camion":                livre_cam or None,
            "livre_chauffeur":             livre_ch or None,
            "restitue_camion":             rest_cam or None,
            "restitue_chauffeur":          rest_ch or None,
            "centre_restitution":          centre or None,
            "date_debut_surestarie":       _date_str(date_debut_sur),
            "date_restitution_estimative": _date_str(date_rest_est),
            "taux_de_change":              taux,
            "nbr_jours_surestarie_estimes": jours_est,
            "nbr_jours_perdu_douane":      jours_douane,
            "nbr_jour_surestarie_facture": jours_facture,
            "montant_facture_check":       facture_check,
            "montant_facture_da":          montant,
            "n_facture_cm":                n_facture or None,
            "date_declaration_douane":     _date_str(date_decl),
            "date_liberation_douane":      _date_str(date_lib),
            "commentaire":                 comment or None,
        })
        st.toast("Saved ✨", icon="✅")
        st.rerun()


# ── PROCESSING ────────────────────────────────────────────────────────────────
elif page == "processing":
    render_page_header("Processing", "Import PDF documents — upload or batch from directory", "⚙️")

    # Ollama status
    try:
        import requests, config as _cfg2
        r = requests.get(_cfg2.OLLAMA_URL.replace("/api/generate",""), timeout=2)
        if r.status_code == 200:
            st.markdown("""<div style="display:inline-flex;align-items:center;gap:8px;
                background:rgba(16,185,129,0.1);border:1px solid rgba(16,185,129,0.3);
                border-radius:8px;padding:6px 14px;margin-bottom:12px;font-size:0.875rem;">
                <span style="color:#10B981;">●</span>
                <span style="color:#E6E9F2;">Ollama is running and ready</span></div>""",
                unsafe_allow_html=True)
        else: raise Exception()
    except Exception:
        st.markdown("""<div style="display:inline-flex;align-items:center;gap:8px;
            background:rgba(245,158,11,0.1);border:1px solid rgba(245,158,11,0.3);
            border-radius:8px;padding:6px 14px;margin-bottom:12px;font-size:0.875rem;">
            <span style="color:#F59E0B;">⚠</span>
            <span style="color:#E6E9F2;">Ollama not detected — start it before processing PDFs</span></div>""",
            unsafe_allow_html=True)

    tab_upload, tab_batch = st.tabs(["📁 Upload Files", "📂 Batch from Directory"])

    with tab_upload:
        uploaded = st.file_uploader("Drop PDF files here", type=["pdf"], accept_multiple_files=True)
        save_copy = st.checkbox("💾 Save copies to input directory", value=True)

        if uploaded and st.button("🚀 Process All", type="primary", use_container_width=True):
            results, progress, status_box = [], st.progress(0), st.empty()
            for i, uf in enumerate(uploaded):
                status_box.info(f"📄 {uf.name} ({i+1}/{len(uploaded)})")
                with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                    tmp.write(uf.read())
                    tmp_path = tmp.name
                if save_copy:
                    dest = os.path.join(st.session_state.settings.get("input_dir","data/input"), uf.name)
                    os.makedirs(os.path.dirname(dest), exist_ok=True)
                    shutil.copy(tmp_path, dest)
                results.append(process_pdf_file(tmp_path, uf.name))
                try: os.unlink(tmp_path)
                except Exception: pass
                progress.progress((i+1)/len(uploaded))
            status_box.empty()

            ok = sum(1 for r in results if r["status"]=="OK")
            tc = sum(r.get("container_count",0) for r in results if r["status"]=="OK")
            if ok == len(results): st.balloons()
            st.success(f"{'🎉' if ok==len(results) else '✅'} {ok}/{len(results)} succeeded — **{tc} containers** imported")

            for r in results:
                icon = "✅" if r["status"]=="OK" else "❌"
                cc = f" · {r.get('container_count',0)} containers" if r["status"]=="OK" else ""
                with st.expander(f"{icon} {r['file']} — {r['status']}{cc}"):
                    if r["status"]=="OK":
                        st.caption(f"Shipment #{r['shipment_id']} · action: **{r['db_action']}**")
                        st.json(r.get("data",{}))
                    else:
                        st.error(r.get("error","Unknown error"))

    with tab_batch:
        input_dir = st.session_state.settings.get("input_dir", "data/input")
        st.info(f"📂 Scanning: `{input_dir}`")
        pdf_files = []
        if os.path.isdir(input_dir):
            pdf_files = sorted(f for f in os.listdir(input_dir) if f.lower().endswith(".pdf"))

        if pdf_files:
            st.markdown(f"**{len(pdf_files)} PDF(s) found:**")
            cols = st.columns(3)
            for i, f in enumerate(pdf_files):
                cols[i%3].markdown(
                    f"<div style='background:rgba(99,102,241,0.08);border:1px solid rgba(99,102,241,0.2);"
                    f"border-radius:8px;padding:8px 12px;margin-bottom:6px;font-size:0.85rem;'>📄 {f}</div>",
                    unsafe_allow_html=True)
        else:
            st.info(f"No PDFs found in `{input_dir}`. Drop files there and refresh.")

        if st.button("▶ Run Batch", type="primary", disabled=not pdf_files, use_container_width=True):
            results, progress, status_box = [], st.progress(0), st.empty()
            for i, f in enumerate(pdf_files):
                status_box.info(f"📄 {f} ({i+1}/{len(pdf_files)})")
                results.append(process_pdf_file(os.path.join(input_dir, f)))
                progress.progress((i+1)/len(pdf_files))
            status_box.empty()
            ok = sum(1 for r in results if r["status"]=="OK")
            tc = sum(r.get("container_count",0) for r in results if r["status"]=="OK")
            if ok==len(results): st.balloons()
            st.success(f"{'🎉' if ok==len(results) else '✅'} {ok}/{len(results)} processed — {tc} containers")
            for r in results:
                icon = "✅" if r["status"]=="OK" else "❌"
                with st.expander(f"{icon} {r['file']} — {r['status']}"):
                    if r["status"]=="OK": st.json(r.get("data",{}))
                    else: st.error(r.get("error","?"))


# ── EXPORT ────────────────────────────────────────────────────────────────────
elif page == "export":
    render_page_header("Export", "Download in target format (49-col XLSX) or custom CSV/Excel", "📤")

    import config as _cfg3
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    df = fetch_containers_view()

    if df.empty:
        st.info("📭 No data. Import PDFs first.")
        st.stop()

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("""
        <div style="background:linear-gradient(135deg,rgba(20,26,46,0.7),rgba(27,35,65,0.5));
                    backdrop-filter:blur(12px);border:1px solid rgba(99,102,241,.2);
                    border-radius:16px;padding:24px;text-align:center;">
            <div style="font-size:3rem;">📗</div>
            <h3 style="margin:8px 0 4px;color:#E6E9F2;">XLSX — Target Format</h3>
            <p style="color:#8C95B3;margin:0 0 4px;font-size:0.85rem;">
                Matches <strong>Containers actifs</strong> — 49 columns exactly.
            </p>
        </div>""", unsafe_allow_html=True)
        st.markdown("")
        if st.button("🔧 Generate XLSX", type="primary", use_container_width=True, key="gen_xlsx"):
            try:
                from utils.xlsx_export import export_xlsx
                path = _cfg3.DB_PATH.replace(".db", f"_containers_actifs_{ts}.xlsx")
                export_xlsx(path)
                st.session_state["_xlsx_path"] = path
                st.toast("XLSX ready!", icon="📗")
            except Exception as e:
                st.error(f"XLSX export failed: {e}")
        if st.session_state.get("_xlsx_path") and os.path.exists(st.session_state["_xlsx_path"]):
            with open(st.session_state["_xlsx_path"],"rb") as f:
                st.download_button("⬇️ Download XLSX", data=f,
                    file_name=os.path.basename(st.session_state["_xlsx_path"]),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)

    with col2:
        st.markdown("""
        <div style="background:linear-gradient(135deg,rgba(20,26,46,0.7),rgba(27,35,65,0.5));
                    backdrop-filter:blur(12px);border:1px solid rgba(99,102,241,.2);
                    border-radius:16px;padding:24px;text-align:center;">
            <div style="font-size:3rem;">📄</div>
            <h3 style="margin:8px 0 4px;color:#E6E9F2;">CSV — Quick Export</h3>
            <p style="color:#8C95B3;margin:0 0 4px;font-size:0.85rem;">
                All {len(df):,} containers, flat CSV for any tool.
            </p>
        </div>""".format(len=len), unsafe_allow_html=True)
        st.markdown("")
        if st.button("🔧 Generate CSV", use_container_width=True, key="gen_csv"):
            path = _cfg3.DB_PATH.replace(".db", f"_export_{ts}.csv")
            export_to_csv(path)
            st.session_state["_csv_path"] = path
            st.toast("CSV ready!", icon="📄")
        if st.session_state.get("_csv_path") and os.path.exists(st.session_state["_csv_path"]):
            with open(st.session_state["_csv_path"],"rb") as f:
                st.download_button("⬇️ Download CSV", data=f,
                    file_name=os.path.basename(st.session_state["_csv_path"]),
                    mime="text/csv", use_container_width=True)

    st.divider()
    # Custom column export
    render_export_page()


# ── SETTINGS + LOGS ───────────────────────────────────────────────────────────
elif page == "settings":
    render_page_header("Settings & Logs", "Configure Ollama, paths, preferences — and browse processing logs", "🛠️")

    tab_settings, tab_logs = st.tabs(["⚙️ Settings", "📄 Logs"])

    with tab_settings:
        render_settings_page()

    with tab_logs:
        files = _get_log_files()
        logs_dir = st.session_state.settings.get("logs_dir","")

        if not files:
            st.info(f"No logs found in `{logs_dir}`")
        else:
            data = []
            for fn in files:
                try:
                    with open(os.path.join(logs_dir, fn), "r", encoding="utf-8") as f:
                        rec = json.load(f)
                    data.append({
                        "File": rec.get("filename", fn),
                        "Time": rec.get("timestamp",""),
                        "OK": "✅" if rec.get("validation_ok") else "❌",
                        "Action": rec.get("db_action","—"),
                        "Error": rec.get("error") or "—",
                        "_fn": fn,
                    })
                except Exception:
                    data.append({"File":fn,"Time":"","OK":"?","Action":"?","Error":"unreadable","_fn":fn})

            ldf = pd.DataFrame(data)
            lc1, lc2, lc3 = st.columns([1,1,2])
            lc1.metric("✅ Success", int((ldf["OK"]=="✅").sum()))
            lc2.metric("❌ Errors",  int((ldf["OK"]=="❌").sum()))
            if lc3.button("🗑 Clear all logs"):
                for fn in files:
                    try: os.remove(os.path.join(logs_dir, fn))
                    except Exception: pass
                st.toast("Logs cleared", icon="🧹")
                st.rerun()

            flt = st.radio("Show", ["All","Success only","Errors only"], horizontal=True)
            disp = ldf.copy()
            if flt=="Success only": disp = disp[disp["OK"]=="✅"]
            elif flt=="Errors only": disp = disp[disp["OK"]=="❌"]
            st.dataframe(disp.drop(columns=["_fn"]), use_container_width=True, hide_index=True, height=360)

            sel = st.selectbox("Inspect log entry", files, key="log_inspect")
            if sel:
                with open(os.path.join(logs_dir, sel),"r",encoding="utf-8") as f:
                    st.json(json.load(f))


# ═══════════════════════════════════════════════════════════════════════════════
# BOTTOM NAVBAR — always last
# ═══════════════════════════════════════════════════════════════════════════════
st.divider()
render_navbar()
