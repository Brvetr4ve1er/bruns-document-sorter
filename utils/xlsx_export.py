"""Export containers to an xlsx file matching the target 'Containers actifs' layout."""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import XLSX_COLUMNS
from database_logic.database import get_connection


def _parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(str(s)[:10], "%Y-%m-%d")
    except Exception:
        return None


def _row_from_container(r: dict) -> list:
    """Map a joined shipment+container SQLite row to the 49 xlsx columns."""
    tan = r.get("tan") or ""
    item = r.get("item_description") or ""
    return [
        "",                                             # (Ne pas modifier) Container
        "",                                             # checksum placeholder
        _parse_date(r.get("modified_at")),              # Modifié le
        r.get("container_number") or "",                # N° Container
        tan,                                            # N° TAN
        item,                                           # Item (N° TAN) (Commande)
        r.get("compagnie_maritime") or "",              # Compagnie maritime
        r.get("port") or "Port d'Alger",                # Port
        r.get("transitaire") or "",                     # Transitaire
        _parse_date(r.get("etd")),                      # Date shipment
        _parse_date(r.get("eta")),                      # Date accostage
        r.get("statut_container") or "En transit",      # Statut Container
        r.get("size") or "",                            # Container size
        _parse_date(r.get("date_livraison")),           # Date livraison
        r.get("site_livraison") or "",                  # Site livraison
        _parse_date(r.get("date_depotement")),          # Date dépotement
        "",                                             # Modifié par
        _parse_date(r.get("modified_at")),              # Modifié le (2)
        _parse_date(r.get("date_debut_surestarie")),    # Date début Surestarie
        _parse_date(r.get("date_restitution_estimative")), # Date restitution estimative
        r.get("nbr_jours_surestarie_estimes") or 0,
        0,                                              # Coût Surestaries Estimé (USD) — computed later
        r.get("nbr_jours_perdu_douane") or 0,
        0,                                              # Coût Surestaries Estimé (DZD)
        0,                                              # Nbr jours restants pour surestarie
        0,                                              # Nbr jours surestarie
        0,                                              # Coût Surestaries Réel (USD)
        0,                                              # Coût Surestaries Réel (DZD)
        _parse_date(r.get("date_restitution")),         # Date réstitution
        r.get("restitue_camion") or "",
        r.get("restitue_chauffeur") or "",
        r.get("centre_restitution") or "",
        0, 0, 0, 0,                                     # 4 Check columns
        _parse_date(r.get("created_at")),               # Créé le
        "",                                             # Créé par
        "",                                             # Check avis d'arrivée-restitution
        r.get("taux_de_change") or 0,                   # Taux de change
        r.get("livre_camion") or "",
        r.get("livre_chauffeur") or "",
        r.get("montant_facture_check") or "No",
        r.get("nbr_jour_surestarie_facture") or 0,
        r.get("montant_facture_da") or 0,
        r.get("n_facture_cm") or "",
        r.get("commentaire") or "",
        _parse_date(r.get("date_declaration_douane")),
        _parse_date(r.get("date_liberation_douane")),
    ]


def export_xlsx(output_path: str) -> str:
    """Write xlsx of all containers in target format. Returns the output path."""
    conn = get_connection()
    rows = conn.execute("""
        SELECT c.*,
               s.tan, s.item_description, s.compagnie_maritime, s.port,
               s.transitaire, s.etd, s.eta
        FROM containers c
        JOIN shipments s ON s.id = c.shipment_id
        ORDER BY s.id, c.id
    """).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Containers actifs"

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for i, col in enumerate(XLSX_COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Rows
    for ridx, r in enumerate(rows, start=2):
        data = _row_from_container(dict(r))
        for cidx, val in enumerate(data, start=1):
            cell = ws.cell(row=ridx, column=cidx, value=val)
            if isinstance(val, datetime):
                cell.number_format = "YYYY-MM-DD"

    # Column widths
    for i, col in enumerate(XLSX_COLUMNS, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = max(14, min(28, len(col) + 2))

    ws.freeze_panes = "A2"
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path
